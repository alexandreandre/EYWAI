"""Lecture Excel tolérante aux exports tiers (Cegid, etc.)."""

from __future__ import annotations

import io
import re
import zipfile
from typing import Any, List
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_PATCH_APPLIED = False


def _apply_openpyxl_compat_patches() -> None:
    """Corrige la faute de frappe biltinId présente dans certains exports Excel."""
    global _PATCH_APPLIED
    if _PATCH_APPLIED or not OPENPYXL_AVAILABLE:
        return
    try:
        from openpyxl.styles.named_styles import _NamedCellStyle
    except ImportError:
        return

    original_init = _NamedCellStyle.__init__

    def patched_init(self, *args: Any, **kwargs: Any) -> None:
        if "biltinId" in kwargs:
            kwargs["builtinId"] = kwargs.pop("biltinId")
        original_init(self, *args, **kwargs)

    _NamedCellStyle.__init__ = patched_init  # type: ignore[method-assign]
    _PATCH_APPLIED = True


def _col_letters_to_index(letters: str) -> int:
    idx = 0
    for ch in letters.upper():
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _cell_value(cell: ET.Element, shared: List[str]) -> str:
    cell_type = cell.get("t")
    inline = cell.find("m:is/m:t", _NS)
    if inline is not None and inline.text is not None:
        return inline.text.strip()
    value = cell.find("m:v", _NS)
    if value is None or value.text is None:
        return ""
    raw = value.text.strip()
    if cell_type == "s":
        try:
            return shared[int(raw)]
        except (ValueError, IndexError):
            return raw
    if isinstance(raw, str) and re.fullmatch(r"-?\d+\.\d+", raw):
        number = float(raw)
        if number.is_integer():
            return str(int(number))
    return raw


def _read_shared_strings(zf: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    strings: List[str] = []
    for item in root.findall("m:si", _NS):
        text_parts = [node.text or "" for node in item.findall(".//m:t", _NS)]
        strings.append("".join(text_parts).strip())
    return strings


def _sheet_xml_path(zf: zipfile.ZipFile) -> str | None:
    names = zf.namelist()
    for candidate in ("xl/worksheets/sheet1.xml",):
        if candidate in names:
            return candidate
    for name in names:
        if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
            return name
    return None


def read_xlsx_raw_rows(content: bytes) -> List[List[str]]:
    """Lit une feuille Excel en lignes brutes, sans dépendre des styles openpyxl."""
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        sheet_path = _sheet_xml_path(zf)
        if not sheet_path:
            return []
        shared = _read_shared_strings(zf)
        root = ET.fromstring(zf.read(sheet_path))
        raw_rows: List[List[str]] = []
        for row_el in root.findall(".//m:sheetData/m:row", _NS):
            cells: dict[int, str] = {}
            for cell in row_el.findall("m:c", _NS):
                ref = cell.get("r", "")
                col_letters = "".join(ch for ch in ref if ch.isalpha())
                if not col_letters:
                    continue
                cells[_col_letters_to_index(col_letters)] = _cell_value(cell, shared)
            if not cells:
                continue
            max_idx = max(cells)
            raw_rows.append([cells.get(i, "") for i in range(max_idx + 1)])
        return raw_rows


def load_workbook_safe(content: bytes, *, sheet_name: str | None = None):
    """Charge un classeur openpyxl avec correctifs ; lève l'erreur d'origine si échec."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl requis pour lire les fichiers Excel.")
    _apply_openpyxl_compat_patches()
    return load_workbook(io.BytesIO(content), read_only=True, data_only=True)


def iter_sheet_rows(
    content: bytes,
    *,
    sheet_name: str | None = None,
) -> List[List[Any]]:
    """Retourne les lignes d'une feuille ; repli XML si openpyxl échoue."""
    if OPENPYXL_AVAILABLE:
        try:
            wb = load_workbook_safe(content)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            rows = [list(row) if row else [] for row in ws.iter_rows(values_only=True)]
            wb.close()
            if rows:
                return rows
        except Exception:
            pass
    return read_xlsx_raw_rows(content)
