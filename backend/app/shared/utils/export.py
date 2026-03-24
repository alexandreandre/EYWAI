"""
Génération de fichiers d'export (XLSX, CSV) et formatage (devise, période).
Partagé entre modules (ex. CSE, exports).
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_xlsx(
    data: List[Dict[str, Any]],
    headers: List[str],
    sheet_name: str = "Export",
) -> bytes:
    """
    Génère un fichier XLSX à partir de données.

    Args:
        data: Liste de dictionnaires représentant les lignes
        headers: Liste des noms de colonnes
        sheet_name: Nom de la feuille Excel

    Returns:
        Contenu du fichier XLSX
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
        )
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, "")
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d")
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_idx, header in enumerate(headers, start=1):
        max_length = max(
            len(str(header)),
            max(
                (len(str(row.get(header, ""))) for row in data),
                default=0,
            ),
        )
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = min(max_length + 2, 50)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def generate_csv(data: List[Dict[str, Any]], headers: List[str]) -> bytes:
    """
    Génère un fichier CSV à partir de données (UTF-8 avec BOM pour Excel).
    """
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue().encode("utf-8-sig")


def format_currency(value: Optional[float]) -> str:
    """Formate un montant en devise française."""
    if value is None:
        return ""
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def format_period(period: str) -> str:
    """Formate une période YYYY-MM en format lisible (ex. Janvier 2024)."""
    try:
        year, month = period.split("-")
        month_names = [
            "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
            "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
        ]
        return f"{month_names[int(month) - 1]} {year}"
    except Exception:
        return period
