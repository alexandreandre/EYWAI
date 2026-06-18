"""Parseur import récap virements CPAM (Excel/CSV)."""

from __future__ import annotations

import csv
import io
import re
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


DEFAULT_COLUMN_ALIASES = {
    "payment_date": ("date", "date virement", "date operation", "date_op"),
    "amount": ("montant", "amount", "credit", "somme"),
    "employee_name": ("libelle", "libellé", "nom", "beneficiaire", "bénéficiaire"),
    "employee_nir": ("nir", "secu", "sécu", "nss"),
    "bank_reference": ("reference", "référence", "ref", "id virement"),
}


def _normalize_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def detect_column_mapping(headers: List[str]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    norm_headers = {_normalize_header(h): h for h in headers if h}
    for field, aliases in DEFAULT_COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in norm_headers:
                mapping[field] = norm_headers[alias]
                break
    return mapping


def _parse_amount(raw: Any) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return round(float(raw), 2)
    s = str(raw).strip().replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def _parse_date(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    if hasattr(raw, "isoformat"):
        return raw.date().isoformat() if hasattr(raw, "date") else raw.isoformat()[:10]
    s = str(raw).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            from datetime import datetime

            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s if re.match(r"\d{4}-\d{2}-\d{2}", s) else None


def _rows_from_xlsx(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl requis pour lire les fichiers Excel.")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], []
    headers = [str(h or "").strip() for h in header_row]
    rows: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows_iter, start=2):
        if not any(row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
    return headers, rows


def _rows_from_csv(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text), delimiter=";")
    if not reader.fieldnames:
        reader = csv.DictReader(io.StringIO(text), delimiter=",")
    headers = list(reader.fieldnames or [])
    return headers, list(reader)


def parse_bank_recap_file(
    filename: str,
    content: bytes,
    column_mapping: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """Parse un fichier récap virements CPAM."""
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xls")):
        headers, raw_rows = _rows_from_xlsx(content)
    else:
        headers, raw_rows = _rows_from_csv(content)

    mapping = column_mapping or detect_column_mapping(headers)
    lines: List[Dict[str, Any]] = []
    anomalies: List[str] = []

    for i, row in enumerate(raw_rows):
        amount_col = mapping.get("amount")
        amt = _parse_amount(row.get(amount_col)) if amount_col else None
        if amt is None or amt <= 0:
            anomalies.append(f"Ligne {i + 1}: montant invalide")
            continue
        date_col = mapping.get("payment_date")
        name_col = mapping.get("employee_name")
        nir_col = mapping.get("employee_nir")
        ref_col = mapping.get("bank_reference")
        lines.append(
            {
                "row_index": i,
                "amount": amt,
                "payment_date": _parse_date(row.get(date_col)) if date_col else None,
                "employee_name_raw": str(row.get(name_col) or "").strip() if name_col else "",
                "employee_nir": str(row.get(nir_col) or "").strip() if nir_col else "",
                "bank_reference": str(row.get(ref_col) or "").strip() if ref_col else "",
                "raw": {k: str(v) if v is not None else "" for k, v in row.items()},
            }
        )

    return {
        "headers": headers,
        "detected_mapping": mapping,
        "lines": lines,
        "anomalies": anomalies,
        "line_count": len(lines),
    }
