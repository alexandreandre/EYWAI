"""Commandes compétences."""

from __future__ import annotations

from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from app.modules.competencies.application import queries
from app.modules.competencies.infrastructure.repository import competencies_repository
from app.modules.competencies.schemas.requests import (
    CompetencyRefCreate,
    CompetencyRefUpdate,
    EmployeeCompetencyCreate,
)
from app.modules.competencies.schemas.responses import CompetencyRef, EmployeeCompetency


def create_competency_ref(company_id: str, data: CompetencyRefCreate) -> CompetencyRef:
    payload: Dict[str, Any] = {
        "name": data.name,
        "category": data.category,
        "description": data.description,
        "required_level": data.required_level,
    }
    row = competencies_repository.create_ref(company_id, payload)
    return queries.competency_ref_from_row(row)


def update_competency_ref(
    ref_id: str, company_id: str, data: CompetencyRefUpdate
) -> CompetencyRef:
    payload = data.model_dump(exclude_unset=True)
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    row = competencies_repository.update_ref(ref_id, company_id, payload)
    return queries.competency_ref_from_row(row)


def archive_competency_ref(ref_id: str, company_id: str) -> None:
    competencies_repository.archive_ref(ref_id, company_id)


def evaluate_employee(
    company_id: str, data: EmployeeCompetencyCreate, evaluated_by: str
) -> EmployeeCompetency:
    if not competencies_repository.get_employee_row(data.employee_id, company_id):
        raise LookupError("Employé non trouvé.")
    cref = competencies_repository.get_ref_by_id(data.competency_id, company_id)
    if not cref or str(cref.get("status") or "") == "archived":
        raise LookupError("Compétence non trouvée ou archivée.")
    payload = {
        "employee_id": data.employee_id,
        "competency_id": data.competency_id,
        "score": data.score,
        "evaluation_date": data.evaluation_date.isoformat(),
        "evaluated_by": evaluated_by,
        "comment": data.comment,
    }
    row = competencies_repository.insert_evaluation(company_id, payload)
    return queries.employee_competency_from_row(row)


def export_matrix_excel_bytes(
    company_id: str,
    service_id: Optional[str] = None,
    category: Optional[str] = None,
) -> tuple[bytes, str]:
    matrix = queries.get_matrix(company_id, service_id=service_id, category=category)
    wb = Workbook()
    ws = wb.active
    ws.title = "Matrice"

    fills = {
        0: PatternFill("solid", fgColor="D9D9D9"),
        1: PatternFill("solid", fgColor="FF6B6B"),
        2: PatternFill("solid", fgColor="FFB347"),
        3: PatternFill("solid", fgColor="C8E6C9"),
        4: PatternFill("solid", fgColor="2E7D32"),
    }
    thick_red = Side(style="thick", color="FF0000")
    gap_border = Border(
        left=thick_red, right=thick_red, top=thick_red, bottom=thick_red
    )

    ws.cell(row=1, column=1, value="Collaborateur")
    for j, c in enumerate(matrix.competencies, start=2):
        ws.cell(row=1, column=j, value=str(c.get("name") or ""))

    score_by: Dict[tuple[str, str], tuple[int, bool, Optional[int]]] = {}
    for cell in matrix.cells:
        score_by[(cell.employee_id, cell.competency_id)] = (
            cell.score,
            cell.is_gap,
            cell.required_level,
        )

    for i, emp in enumerate(matrix.employees, start=2):
        eid = str(emp["id"])
        ws.cell(row=i, column=1, value=str(emp.get("name") or ""))
        for j, c in enumerate(matrix.competencies, start=2):
            cid = str(c["id"])
            score, is_gap, req = score_by.get((eid, cid), (0, False, c.get("required_level")))
            cell = ws.cell(row=i, column=j, value="—" if score == 0 else str(score))
            cell.fill = fills.get(int(score), fills[0])
            if int(score) == 4:
                cell.font = Font(color="FFFFFF", bold=True)
            if is_gap:
                cell.border = gap_border

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    d = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    fname = f"matrice_competences_{company_id}_{d}.xlsx"
    return bio.getvalue(), fname
