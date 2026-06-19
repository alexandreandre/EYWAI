"""Parseur tabulaire générique CSV/XLSX pour import pointages."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


DEFAULT_COLUMN_ALIASES = {
    "matricule": (
        "matricule",
        "mat",
        "id",
        "badge",
        "numero",
        "numéro",
        "time_tracking_id",
    ),
    "last_name": ("nom", "lastname", "name", "salarié", "salarié(e)"),
    "first_name": ("prenom", "prénom", "firstname"),
    "date": ("date", "jour", "day", "date_pointage"),
    "hours": (
        "heures",
        "hours",
        "duree",
        "durée",
        "temps",
        "total",
        "heures_faites",
    ),
}


@dataclass
class TabularDayRow:
    matricule: Optional[str] = None
    raw_name: Optional[str] = None
    jour: int = 0
    month: int = 0
    year: int = 0
    heures: float = 0.0
    raw_payload: Dict[str, Any] = field(default_factory=dict)


@dataclass
class TabularParseResult:
    confidence: float = 0.0
    rows: List[TabularDayRow] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    headers: List[str] = field(default_factory=list)
    column_mapping: Dict[str, str] = field(default_factory=dict)


def _normalize_header(h: str) -> str:
    text = re.sub(r"\s+", " ", (h or "").strip().lower())
    text = text.replace("é", "e").replace("è", "e").replace("ê", "e")
    return text


def detect_column_mapping(headers: List[str]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    norm = {_normalize_header(h): h for h in headers if h}
    for field_name, aliases in DEFAULT_COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in norm:
                mapping[field_name] = norm[alias]
                break
    return mapping


def _rows_from_csv(content: bytes, delimiter: str | None = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = content.decode("utf-8", errors="replace")

    if delimiter is None:
        sample = text[:2048]
        delimiter = ";" if sample.count(";") > sample.count(",") else ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = list(reader.fieldnames or [])
    return headers, list(reader)


def _rows_from_xlsx(content: bytes, sheet_name: str | None = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl requis pour lire les fichiers Excel.")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], []
    headers = [str(h or "").strip() for h in header_row]
    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        if not any(row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
    return headers, rows


def read_tabular_preview(
    content: bytes,
    filename: str,
    *,
    max_rows: int = 5,
    options: Optional[Dict[str, Any]] = None,
) -> Tuple[List[str], List[Dict[str, Any]], Dict[str, str]]:
    opts = options or {}
    lower = (filename or "").lower()
    skip_rows = int(opts.get("skip_rows") or 0)
    if lower.endswith((".xlsx", ".xls")):
        headers, rows = _rows_from_xlsx(content, opts.get("sheet_name"))
    else:
        headers, rows = _rows_from_csv(content, opts.get("delimiter"))
    if skip_rows:
        rows = rows[skip_rows:]
    mapping = detect_column_mapping(headers)
    return headers, rows[:max_rows], mapping


def _parse_hours(raw: Any, *, decimal_separator: str = ".") -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return round(float(raw), 2)
    s = str(raw).strip().lower().replace(" ", "")
    if not s or s in ("-", "—"):
        return None
    if decimal_separator == ",":
        s = s.replace(",", ".")
    m = re.match(r"^(\d+)[:h.](\d{2})$", s)
    if m:
        return round(int(m.group(1)) + int(m.group(2)) / 60.0, 2)
    m = re.match(r"^(\d{1,2})h(\d{2})?$", s)
    if m:
        mins = int(m.group(2) or 0)
        return round(int(m.group(1)) + mins / 60.0, 2)
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return None
    try:
        val = float(s)
        return round(val, 2) if 0 <= val <= 24 else None
    except ValueError:
        return None


def _parse_date_value(
    raw: Any,
    *,
    date_format: str | None,
    default_year: int,
    default_month: int,
) -> Optional[Tuple[int, int, int]]:
    if raw is None:
        return None
    if hasattr(raw, "date"):
        d = raw.date() if hasattr(raw, "date") else raw
        return d.year, d.month, d.day
    if isinstance(raw, (int, float)) and raw > 30000:
        base = date(1899, 12, 30)
        d = base + timedelta(days=int(raw))
        return d.year, d.month, d.day
    s = str(raw).strip()
    formats = [date_format] if date_format else []
    formats.extend(["%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%Y"])
    for fmt in formats:
        if not fmt:
            continue
        try:
            d = datetime.strptime(s[:10], fmt).date()
            return d.year, d.month, d.day
        except ValueError:
            continue
    m = re.match(r"^(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?$", s)
    if m:
        day, mon = int(m.group(1)), int(m.group(2))
        y_raw = m.group(3)
        year = default_year
        if y_raw:
            year = int(y_raw)
            if year < 100:
                year += 2000
        return year, mon, day
    if s.isdigit() and 1 <= int(s) <= 31:
        return default_year, default_month, int(s)
    return None


def parse_tabular_file(
    content: bytes,
    filename: str,
    *,
    target_year: int,
    target_month: int,
    column_mapping: Optional[Dict[str, str]] = None,
    options: Optional[Dict[str, Any]] = None,
) -> TabularParseResult:
    opts = options or {}
    lower = (filename or "").lower()
    skip_rows = int(opts.get("skip_rows") or 0)
    decimal_sep = str(opts.get("decimal_separator") or ".")

    if lower.endswith((".xlsx", ".xls")):
        headers, raw_rows = _rows_from_xlsx(content, opts.get("sheet_name"))
    else:
        headers, raw_rows = _rows_from_csv(content, opts.get("delimiter"))

    if skip_rows:
        raw_rows = raw_rows[skip_rows:]

    mapping = column_mapping or detect_column_mapping(headers)
    result = TabularParseResult(headers=headers, column_mapping=mapping)

    if not mapping.get("hours") and not mapping.get("date"):
        result.warnings.append(
            "Colonnes heures ou date non détectées — vérifiez le mapping."
        )
        return result

    date_col = mapping.get("date")
    hours_col = mapping.get("hours")
    mat_col = mapping.get("matricule")
    last_col = mapping.get("last_name")
    first_col = mapping.get("first_name")

    parsed_count = 0
    for idx, row in enumerate(raw_rows):
        heures = (
            _parse_hours(row.get(hours_col), decimal_separator=decimal_sep)
            if hours_col
            else None
        )
        date_parts = (
            _parse_date_value(
                row.get(date_col),
                date_format=opts.get("date_format"),
                default_year=target_year,
                default_month=target_month,
            )
            if date_col
            else None
        )
        if heures is None and date_parts is None:
            continue
        if date_parts is None:
            result.warnings.append(f"Ligne {idx + 1} : date illisible.")
            continue
        y, m, d = date_parts
        if m != target_month or y != target_year:
            continue
        if heures is None:
            heures = 0.0

        mat = str(row.get(mat_col) or "").strip() if mat_col else None
        name_parts = []
        if last_col and row.get(last_col):
            name_parts.append(str(row.get(last_col)).strip())
        if first_col and row.get(first_col):
            name_parts.append(str(row.get(first_col)).strip())
        raw_name = " ".join(name_parts) if name_parts else None

        result.rows.append(
            TabularDayRow(
                matricule=mat or None,
                raw_name=raw_name,
                jour=d,
                month=m,
                year=y,
                heures=heures,
                raw_payload=dict(row),
            )
        )
        parsed_count += 1

    if parsed_count == 0:
        result.confidence = 0.0
        result.warnings.append("Aucune ligne exploitable pour la période cible.")
    elif parsed_count >= 5:
        result.confidence = 0.9
    else:
        result.confidence = 0.65

    return result


__all__ = [
    "TabularDayRow",
    "TabularParseResult",
    "detect_column_mapping",
    "parse_tabular_file",
    "read_tabular_preview",
]
