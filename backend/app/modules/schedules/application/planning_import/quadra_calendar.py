"""Parseur calendrier prévu Excel Quadra (feuille par salarié, blocs mensuels)."""

from __future__ import annotations

import calendar as cal_mod
import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from typing import Any, Dict, List, Sequence, Tuple

from app.modules.schedules.application.employee_match import (
    rank_planning_sheet_candidates,
    resolve_employee_for_planning_sheet,
)
from app.modules.schedules.application.timesheet_import.tabular_period import (
    ImportPeriodConfig,
)
from app.modules.schedules.schemas.ai import AiDayEntry, RosterEmployee

SKIP_SHEETS = frozenset({"SOMMAIRE", "MODELE"})

MONTH_FR: dict[str, int] = {
    "JANVIER": 1,
    "FEVRIER": 2,
    "MARS": 3,
    "AVRIL": 4,
    "MAI": 5,
    "JUIN": 6,
    "JUILLET": 7,
    "AOUT": 8,
    "SEPTEMBRE": 9,
    "OCTOBRE": 10,
    "NOVEMBRE": 11,
    "DECEMBRE": 12,
}

DEFAULT_DAILY_HOURS = 7.8


def _normalize_text(value: str | None) -> str:
    folded = unicodedata.normalize("NFD", value or "")
    stripped = "".join(c for c in folded if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", stripped.upper().strip())


def _sheet_key(name: str) -> str:
    return _normalize_text(name)


def _parse_month_header(value: object) -> int | None:
    if value is None:
        return None
    return MONTH_FR.get(_normalize_text(str(value)))


def _is_holiday_label(label: str) -> bool:
    if not label:
        return False
    keywords = (
        "JOUR DE L'AN",
        "PAQUES",
        "PÂQUES",
        "FETE DU W",
        "FETE DU TRAVAIL",
        "ASCENSION",
        "PENTECOTE",
        "PENTECÔTE",
        "PENTE",
        "VICTOIRE",
        "NATIONALE",
        "ASSOMPTION",
        "TOUSSAINT",
        "ARMISTICE",
        "NOEL",
        "NOËL",
    )
    return any(k in label for k in keywords)


def _classify_absence_label(label: str) -> str:
    if _is_holiday_label(label):
        return "ferie"
    if any(k in label for k in ("MALADIE", "CARENCE", "ARRET", "ARRÊT", " AT ", "FIN CDI")):
        return "arret_maladie"
    if any(
        k in label
        for k in (
            "EVF",
            "DECES",
            "DÉCÈS",
            "PATERNIT",
            "MATERNIT",
            "NAISSANCE",
            "PACS",
        )
    ):
        return "conge"
    if any(k in label for k in ("RECUP", "RÉCUP", "REPOS", " HR")):
        return "conge"
    if "JC" in label or "JOUR CADRE" in label:
        return "conge"
    if "FORMATION" in label or label.startswith("SST"):
        return "travail"
    if label.startswith("VM"):
        return "travail"
    if label in ("", " "):
        return "conge"
    return "conge"


def _parse_cp_value(raw: object) -> float | str | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text:
        return None
    upper = _normalize_text(text)
    if upper in ("JFNP", "PENTE", "PENTECOTE", "PENTECÔTE"):
        return upper
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return upper


def _day_entry(
    *,
    jour: int,
    day_type: str,
    heures: float | None,
    note: str | None = None,
) -> dict[str, Any]:
    entry: dict[str, Any] = {
        "jour": jour,
        "type": day_type,
        "heures_prevues": heures,
    }
    if note and day_type == "arret_maladie":
        entry["arret_type"] = "maladie_simple"
    return entry


def classify_planned_day(
    day_date: date,
    hab_raw: object,
    cp_raw: object,
    *,
    daily_hours: float = DEFAULT_DAILY_HOURS,
) -> dict[str, Any]:
    """Convertit une ligne Excel (H.Abs / CP) en entrée calendrier_prevu."""
    jour = day_date.day
    if day_date.weekday() >= 5:
        return _day_entry(jour=jour, day_type="weekend", heures=0.0)

    cp = _parse_cp_value(cp_raw)
    hab_text = ""
    hab_is_blank_marker = False
    if hab_raw is not None and not isinstance(hab_raw, (int, float)):
        raw_str = str(hab_raw)
        if raw_str.strip() == "" and raw_str != "":
            hab_is_blank_marker = True
        else:
            hab_text = raw_str.strip()
    hab_norm = _normalize_text(hab_text)

    if isinstance(hab_raw, (int, float)) and float(hab_raw) < 0:
        return _day_entry(jour=jour, day_type="conge", heures=daily_hours)

    if hab_is_blank_marker:
        return _day_entry(jour=jour, day_type="conge", heures=daily_hours)

    if cp is not None:
        if isinstance(cp, str):
            if cp in ("JFNP", "PENTE", "PENTECOTE", "PENTECÔTE"):
                return _day_entry(jour=jour, day_type="ferie", heures=None)
        elif isinstance(cp, float):
            if cp <= 0:
                pass
            elif cp >= 1:
                return _day_entry(jour=jour, day_type="conge", heures=daily_hours)
            else:
                return _day_entry(
                    jour=jour,
                    day_type="conge",
                    heures=round(daily_hours * cp, 2),
                )

    if hab_text and hab_norm:
        if _is_holiday_label(hab_norm):
            return _day_entry(jour=jour, day_type="ferie", heures=None)
        day_type = _classify_absence_label(hab_norm)
        if day_type == "travail":
            return _day_entry(jour=jour, day_type="travail", heures=daily_hours)
        if day_type == "ferie":
            return _day_entry(jour=jour, day_type="ferie", heures=None)
        if day_type == "arret_maladie":
            return _day_entry(
                jour=jour,
                day_type="arret_maladie",
                heures=daily_hours,
                note=hab_text,
            )
        return _day_entry(jour=jour, day_type="conge", heures=daily_hours)

    return _day_entry(jour=jour, day_type="travail", heures=daily_hours)


def build_default_month_calendar(
    year: int,
    month: int,
    *,
    daily_hours: float = DEFAULT_DAILY_HOURS,
    holiday_days: set[int] | None = None,
) -> list[dict[str, Any]]:
    """Semaine type + week-ends + fériés entreprise."""
    _, n_days = cal_mod.monthrange(year, month)
    holidays = holiday_days or set()
    entries: list[dict[str, Any]] = []
    for jour in range(1, n_days + 1):
        wd = date(year, month, jour).weekday()
        if wd >= 5:
            entries.append(_day_entry(jour=jour, day_type="weekend", heures=0.0))
        elif jour in holidays:
            entries.append(_day_entry(jour=jour, day_type="ferie", heures=None))
        else:
            entries.append(
                _day_entry(jour=jour, day_type="travail", heures=daily_hours)
            )
    return entries


def _header_rows_snapshot(ws, *, max_rows: int = 6) -> list[list[Any]]:
    """Lit les premières lignes du sheet en une passe (compatible read_only)."""
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        rows.append(list(row))
        if len(rows) >= max_rows:
            break
    return rows


def _parse_sheet_month_blocks_from_rows(
    header_rows: list[list[Any]],
) -> list[tuple[int, int, int]]:
    """Retourne [(month, base_col_1based, header_row_1based)] à partir des premières lignes."""
    for row_idx, values in enumerate(header_rows, start=1):
        row_blocks: list[tuple[int, int, int]] = []
        for col_idx in range(0, len(values), 4):
            month = _parse_month_header(values[col_idx])
            if month:
                row_blocks.append((month, col_idx + 1, row_idx))
        if row_blocks:
            return row_blocks
    return []


def _parse_sheet_month_blocks(ws) -> list[tuple[int, int, int]]:
    return _parse_sheet_month_blocks_from_rows(_header_rows_snapshot(ws))


def _parse_employee_sheet_from_rows(
    day_rows: list[list[Any]],
    blocks: list[tuple[int, int, int]],
    *,
    year: int,
    daily_hours: float,
) -> dict[int, list[dict[str, Any]]]:
    by_month: dict[int, list[dict[str, Any]]] = {}
    for month, base_col, _header_row in blocks:
        entries: list[dict[str, Any]] = []
        day_col = base_col  # base_col+1 1-based → index base_col in 0-based of slice
        hab_col = base_col + 1
        cp_col = base_col + 2
        for values in day_rows:
            if day_col >= len(values):
                continue
            day_raw = values[day_col]
            if day_raw is None:
                continue
            try:
                jour = int(day_raw)
            except (TypeError, ValueError):
                continue
            try:
                day_date = date(year, month, jour)
            except ValueError:
                continue
            hab = values[hab_col] if hab_col < len(values) else None
            cp = values[cp_col] if cp_col < len(values) else None
            entries.append(classify_planned_day(day_date, hab, cp, daily_hours=daily_hours))
        if entries:
            by_month[month] = sorted(entries, key=lambda e: e["jour"])
    return by_month


def parse_employee_sheet(
    ws,
    *,
    year: int,
    daily_hours: float = DEFAULT_DAILY_HOURS,
) -> dict[int, list[dict[str, Any]]]:
    header_rows = _header_rows_snapshot(ws)
    blocks = _parse_sheet_month_blocks_from_rows(header_rows)
    if not blocks:
        return {}
    header_row = blocks[0][2]
    day_rows = list(
        ws.iter_rows(
            min_row=header_row + 1,
            max_row=header_row + 31,
            values_only=True,
        )
    )
    day_rows = [list(r) for r in day_rows]
    return _parse_employee_sheet_from_rows(
        day_rows, blocks, year=year, daily_hours=daily_hours
    )


def _load_workbook_read_only(content: bytes):
    import openpyxl

    return openpyxl.load_workbook(
        io.BytesIO(content), data_only=True, read_only=True
    )


def is_quadra_planning_workbook(content: bytes, filename: str) -> bool:
    lower = (filename or "").lower()
    if not lower.endswith((".xlsx", ".xls")):
        return False
    try:
        wb = _load_workbook_read_only(content)
    except Exception:
        return False
    try:
        sheet_keys = {_sheet_key(name) for name in wb.sheetnames}
        if "MODELE" in sheet_keys:
            return True
        for name in wb.sheetnames:
            key = _sheet_key(name)
            if key in SKIP_SHEETS:
                continue
            ws = wb[name]
            if _parse_sheet_month_blocks(ws):
                return True
        return False
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _iter_target_months(config: ImportPeriodConfig) -> set[tuple[int, int]] | None:
    if config.mode == "auto":
        return None
    if config.mode == "month":
        return {(config.year, config.month)}
    if config.mode == "year":
        return {(config.year, month) for month in range(1, 13)}
    if config.mode == "range":
        sy = config.start_year or config.year
        sm = config.start_month or 1
        ey = config.end_year or config.year
        em = config.end_month or 12
        months: set[tuple[int, int]] = set()
        y, m = sy, sm
        while (y, m) <= (ey, em):
            months.add((y, m))
            if m == 12:
                y += 1
                m = 1
            else:
                m += 1
        return months
    return None


def _entry_to_ai_day(entry: dict[str, Any]) -> AiDayEntry:
    return AiDayEntry(
        jour=int(entry["jour"]),
        heures=entry.get("heures_prevues"),
        type=str(entry.get("type") or "travail"),
        nature="prevu",
    )


@dataclass
class QuadraPlanningParseResult:
    year: int
    month: int
    month_groups: List[Dict[str, Any]] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    sheets_parsed: int = 0
    sheets_unmatched: List[str] = field(default_factory=list)


def _build_sommaire_hints(wb, employee_sheet_keys: set[str]) -> dict[str, str]:
    """Associe une clé de feuille salarié à un libellé complet lu dans Sommaire."""
    hints: dict[str, str] = {}
    sommaire_ws = None
    for name in wb.sheetnames:
        if _sheet_key(name) == "SOMMAIRE":
            sommaire_ws = wb[name]
            break
    if not sommaire_ws:
        return hints

    key_cache: list[tuple[str, str, set[str]]] = []
    for key in employee_sheet_keys:
        key_norm = _normalize_text(key)
        key_cache.append((key, key_norm, set(key_norm.split())))

    for row_values in sommaire_ws.iter_rows(min_row=1, max_col=2, values_only=True):
        last = str(row_values[0] or "").strip() if len(row_values) > 0 else ""
        first = str(row_values[1] or "").strip() if len(row_values) > 1 else ""
        full = " ".join(part for part in (last, first) if part).strip()
        if not full or len(full.split()) < 2:
            continue
        full_norm = _normalize_text(full)
        full_tokens = set(full_norm.split())
        best_key: str | None = None
        best_score = 0
        for key, key_norm, key_tokens in key_cache:
            if key_norm and key_norm in full_norm:
                score = len(key_norm) + 20
            elif key_tokens and key_tokens.issubset(full_tokens):
                score = len(key_tokens) * 10
            else:
                continue
            if score > best_score:
                best_key = key
                best_score = score
        if best_key:
            hints[best_key] = full
    return hints


def _extract_sheet_employee_hint_from_rows(
    header_rows: list[list[Any]], sheet_name: str
) -> str | None:
    sheet_key = _sheet_key(sheet_name)
    if not sheet_key:
        return None

    best: str | None = None
    best_score = 0
    for row_idx, values in enumerate(header_rows[:3], start=1):
        for col_idx, raw in enumerate(values[:16]):
            if raw is None:
                continue
            text = str(raw).strip()
            if not text or len(text.split()) < 2:
                continue
            norm = _normalize_text(text)
            if sheet_key not in norm:
                continue
            score = len(sheet_key)
            if norm != sheet_key:
                score += 30
            if row_idx == 1:
                score += 5
            if score > best_score:
                best = text
                best_score = score
    return best


def _extract_sheet_employee_hint(ws, sheet_name: str) -> str | None:
    return _extract_sheet_employee_hint_from_rows(
        _header_rows_snapshot(ws), sheet_name
    )


def parse_quadra_planning_workbook(
    content: bytes,
    filename: str,
    *,
    year: int,
    period_config: ImportPeriodConfig,
    roster: Sequence[RosterEmployee],
    daily_hours: float = DEFAULT_DAILY_HOURS,
) -> QuadraPlanningParseResult:
    wb = _load_workbook_read_only(content)
    try:
        target_months = _iter_target_months(period_config)
        warnings: List[str] = []
        groups_map: Dict[Tuple[int, int], List[Dict[str, Any]]] = {}
        sheets_parsed = 0
        sheets_unmatched: List[str] = []
        used_employee_ids: set[str] = set()
        roster_list = list(roster)

        employee_sheets: List[tuple[str, list[list[Any]], list[tuple[int, int, int]]]] = []
        for sheet_name in wb.sheetnames:
            sheet_key = _sheet_key(sheet_name)
            if sheet_key in SKIP_SHEETS:
                continue
            ws = wb[sheet_name]
            header_rows = _header_rows_snapshot(ws)
            blocks = _parse_sheet_month_blocks_from_rows(header_rows)
            if blocks:
                employee_sheets.append((sheet_name, header_rows, blocks))

        sommaire_hints = _build_sommaire_hints(
            wb, {_sheet_key(name) for name, _, _ in employee_sheets}
        )

        for sheet_name, header_rows, blocks in sorted(
            employee_sheets, key=lambda item: _sheet_key(item[0])
        ):
            sheet_key = _sheet_key(sheet_name)
            sheets_parsed += 1
            hint_name = _extract_sheet_employee_hint_from_rows(
                header_rows, sheet_name
            ) or sommaire_hints.get(sheet_key)
            ranked = rank_planning_sheet_candidates(
                sheet_name.strip(),
                roster_list,
                exclude_employee_ids=used_employee_ids,
                hint_name=hint_name,
                limit=8,
            )
            suggestions = ranked[:5]
            match = resolve_employee_for_planning_sheet(
                sheet_name.strip(),
                roster_list,
                exclude_employee_ids=used_employee_ids,
                hint_name=hint_name,
                precomputed_candidates=ranked,
            )
            if match.employee_id and match.review_status in ("ok", "warning"):
                used_employee_ids.add(str(match.employee_id))

            header_row = blocks[0][2]
            ws = wb[sheet_name]
            day_rows = [
                list(r)
                for r in ws.iter_rows(
                    min_row=header_row + 1,
                    max_row=header_row + 31,
                    values_only=True,
                )
            ]
            parsed = _parse_employee_sheet_from_rows(
                day_rows, blocks, year=year, daily_hours=daily_hours
            )

            for month_num, entries in parsed.items():
                if target_months is not None and (year, month_num) not in target_months:
                    continue
                days = [_entry_to_ai_day(entry) for entry in entries]
                if not days:
                    continue
                groups_map.setdefault((year, month_num), []).append(
                    {
                        "employee_id": match.employee_id,
                        "raw_name": match.raw_name or sheet_name.strip(),
                        "matched_name": match.matched_name,
                        "time_tracking_id": match.time_tracking_id,
                        "review_status": match.review_status,
                        "match_method": match.match_method,
                        "match_confidence": match.match_confidence,
                        "suggested_employee_ids": [e.id for e in suggestions],
                        "sommaire_hint": hint_name,
                        "days": [d.model_dump(mode="json") for d in days],
                    }
                )
                if (
                    match.review_status == "error"
                    and sheet_name.strip() not in sheets_unmatched
                ):
                    sheets_unmatched.append(sheet_name.strip())
                for w in match.warnings:
                    if w not in warnings:
                        warnings.append(w)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if sheets_parsed == 0:
        raise ValueError(
            "Aucune feuille calendrier Quadra reconnue (attendu : blocs mensuels JANVIER…)."
        )
    if not groups_map:
        raise ValueError(
            "Aucune donnée calendrier dans la période demandée — vérifiez l'année ou la plage."
        )

    month_groups = [
        {"year": y, "month": m, "employees": employees}
        for (y, m), employees in sorted(groups_map.items())
    ]
    anchor_year, anchor_month = month_groups[0]["year"], month_groups[0]["month"]
    if sheets_unmatched:
        warnings.append(
            f"{len(sheets_unmatched)} feuille(s) sans salarié rapproché : "
            + ", ".join(sheets_unmatched[:8])
            + ("…" if len(sheets_unmatched) > 8 else "")
        )

    return QuadraPlanningParseResult(
        year=anchor_year,
        month=anchor_month,
        month_groups=month_groups,
        warnings=warnings,
        sheets_parsed=sheets_parsed,
        sheets_unmatched=sheets_unmatched,
    )


__all__ = [
    "DEFAULT_DAILY_HOURS",
    "build_default_month_calendar",
    "classify_planned_day",
    "is_quadra_planning_workbook",
    "parse_employee_sheet",
    "parse_quadra_planning_workbook",
]
