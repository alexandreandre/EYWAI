"""Lecture Excel/CSV et détection colonnes pour import RIB."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


RIB_COLUMN_ALIASES = (
    "rib",
    "iban",
    "releve identite bancaire",
    "relevé identité bancaire",
    "coordonnees bancaires",
    "coordonnées bancaires",
)

BIC_COLUMN_ALIASES = ("bic", "swift", "code bic")

MATRICULE_ALIASES = (
    "matricule",
    "mat",
    "badge",
    "numero",
    "numéro",
    "time_tracking_id",
)

LAST_NAME_ALIASES = ("nom", "lastname", "name", "salarié", "salarié(e)", "salarié(e)s")
FIRST_NAME_ALIASES = ("prenom", "prénom", "firstname")
FULL_NAME_ALIASES = ("nom prenom", "nom prénom", "nom et prenom", "nom complet", "identite", "identité")
EMAIL_ALIASES = ("email", "e-mail", "mail", "courriel")


@dataclass
class TabularSheet:
    headers: List[str] = field(default_factory=list)
    rows: List[Dict[str, str]] = field(default_factory=list)


def _normalize_header(h: str) -> str:
    text = re.sub(r"\s+", " ", (h or "").strip().lower())
    text = text.replace("é", "e").replace("è", "e").replace("ê", "e").replace("à", "a")
    return text


def _match_alias(header: str, aliases: Tuple[str, ...]) -> bool:
    norm = _normalize_header(header)
    return norm in aliases or any(alias in norm for alias in aliases if len(alias) >= 3)


def detect_rib_column_mapping(headers: List[str]) -> Dict[str, str]:
    """Retourne mapping logique -> nom de colonne source."""
    mapping: Dict[str, str] = {}
    for header in headers:
        if not header:
            continue
        if "rib" not in mapping and _match_alias(header, RIB_COLUMN_ALIASES):
            mapping["rib"] = header
        elif "bic" not in mapping and _match_alias(header, BIC_COLUMN_ALIASES):
            mapping["bic"] = header
        elif "matricule" not in mapping and _match_alias(header, MATRICULE_ALIASES):
            mapping["matricule"] = header
        elif "last_name" not in mapping and _match_alias(header, LAST_NAME_ALIASES):
            mapping["last_name"] = header
        elif "first_name" not in mapping and _match_alias(header, FIRST_NAME_ALIASES):
            mapping["first_name"] = header
        elif "full_name" not in mapping and _match_alias(header, FULL_NAME_ALIASES):
            mapping["full_name"] = header
        elif "email" not in mapping and _match_alias(header, EMAIL_ALIASES):
            mapping["email"] = header
    return mapping


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_csv(content: bytes) -> TabularSheet:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        return TabularSheet()
    headers = [str(h).strip() for h in all_rows[0]]
    data: List[Dict[str, str]] = []
    for row in all_rows[1:]:
        if not any(cell.strip() for cell in row if cell):
            continue
        payload = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            payload[header] = row[idx].strip() if idx < len(row) else ""
        data.append(payload)
    return TabularSheet(headers=headers, rows=data)


def _read_xlsx(content: bytes) -> TabularSheet:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl requis pour lire les fichiers Excel.")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return TabularSheet()
    headers = [_cell_str(h) for h in header_row]
    data: List[Dict[str, str]] = []
    for row in rows_iter:
        if not row or not any(v is not None and str(v).strip() for v in row):
            continue
        payload = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            payload[header] = _cell_str(row[idx]) if idx < len(row) else ""
        data.append(payload)
    wb.close()
    return TabularSheet(headers=headers, rows=data)


def read_tabular_file(content: bytes, filename: str) -> TabularSheet:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return _read_csv(content)
    if lower.endswith((".xlsx", ".xls")):
        return _read_xlsx(content)
    raise ValueError("Format non supporté. Utilisez un fichier Excel (.xlsx) ou CSV.")


def row_value(row: Dict[str, str], column: Optional[str]) -> str:
    if not column:
        return ""
    return (row.get(column) or "").strip()
