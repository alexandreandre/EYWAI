# Export « Paiement organismes » — échéances URSSAF, retraite, mutuelle, prévoyance.
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.modules.exports.infrastructure.export_charges_sociales import (
    get_charges_sociales_data,
)
from app.shared.utils.export import format_period, generate_csv

HEADERS = [
    "Organisme",
    "Libellé cotisation",
    "Part salariale",
    "Part patronale",
    "Total à payer",
    "Échéance",
    "Référence paiement",
    "IBAN organisme",
]


def _build_payment_rows(
    detail_rows: List[Dict[str, Any]],
    period: str,
) -> List[Dict[str, Any]]:
    year, month = map(int, period.split("-"))
    echeance = f"{year}-{month:02d}-15"
    rows: List[Dict[str, Any]] = []
    for item in detail_rows:
        total = float(item.get("Total cotisations", 0) or 0)
        if total <= 0:
            continue
        org = item.get("Organisme", "AUTRE")
        rows.append(
            {
                "Organisme": org,
                "Libellé cotisation": item.get("Libellé cotisation", ""),
                "Part salariale": item.get("Part salariale", 0),
                "Part patronale": item.get("Part patronale", 0),
                "Total à payer": round(total, 2),
                "Échéance": echeance,
                "Référence paiement": f"ORG-{period}-{org}",
                "IBAN organisme": "",
            }
        )
    return rows


def preview_paiement_organismes(
    company_id: str,
    period: str,
    employee_ids: Optional[List[str]] = None,
) -> Dict[str, Any]:
    detail_rows, summary_rows, totals = get_charges_sociales_data(
        company_id, period, employee_ids
    )
    payment_rows = _build_payment_rows(detail_rows, period)
    warnings: List[str] = []
    if not payment_rows:
        warnings.append("Aucune cotisation à payer pour cette période.")
    return {
        "employees_count": totals.get("employees_count", 0),
        "totals": {
            **totals,
            "operations_count": len(payment_rows),
            "total_amount": sum(r["Total à payer"] for r in payment_rows),
        },
        "anomalies": [],
        "warnings": warnings,
        "can_generate": True,
        "details": {"lines": payment_rows, "organismes": summary_rows},
    }


def generate_paiement_organismes_export(
    company_id: str,
    period: str,
    employee_ids: Optional[List[str]] = None,
    file_format: str = "xlsx",
) -> bytes:
    detail_rows, _, _ = get_charges_sociales_data(
        company_id, period, employee_ids
    )
    rows = _build_payment_rows(detail_rows, period)
    if file_format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Paiement organismes"
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col_idx, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, h in enumerate(HEADERS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    return generate_csv(rows, HEADERS)
