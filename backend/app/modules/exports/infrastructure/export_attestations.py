# Export « Attestations & annexes » — attestations employeur et salaire.
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

from openpyxl import Workbook

from app.core.database import supabase
from app.shared.utils.export import format_period, generate_csv

HEADERS = [
    "Employé",
    "Matricule",
    "Type attestation",
    "Période",
    "Brut",
    "Net imposable",
    "Net à payer",
    "Statut",
]


def _fetch_attestation_rows(
    company_id: str,
    period: str,
    employee_ids: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    year, month = map(int, period.split("-"))
    query = (
        supabase.table("payslips")
        .select(
            """
            id, employee_id, payslip_data,
            employees!inner(id, first_name, last_name, company_id)
            """
        )
        .eq("company_id", company_id)
        .eq("year", year)
        .eq("month", month)
    )
    if employee_ids:
        query = query.in_("employee_id", employee_ids)
    response = query.execute()
    rows: List[Dict[str, Any]] = []
    for ps in response.data or []:
        emp = ps.get("employees") or {}
        pd = ps.get("payslip_data") or {}
        if not isinstance(pd, dict):
            continue
        name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
        synthese = pd.get("synthese_net") or {}
        net_imp = float(synthese.get("net_imposable", 0) or 0) if isinstance(synthese, dict) else 0
        rows.append(
            {
                "Employé": name,
                "Matricule": str(emp.get("id", ""))[:8],
                "Type attestation": "Attestation de salaire",
                "Période": format_period(period),
                "Brut": float(pd.get("salaire_brut", 0) or 0),
                "Net imposable": net_imp,
                "Net à payer": float(pd.get("net_a_payer", 0) or 0),
                "Statut": "Généré",
            }
        )
    return rows


def preview_attestations(
    company_id: str,
    period: str,
    employee_ids: Optional[List[str]] = None,
) -> Dict[str, Any]:
    rows = _fetch_attestation_rows(company_id, period, employee_ids)
    warnings: List[str] = []
    if not rows:
        warnings.append("Aucun bulletin trouvé pour générer les attestations.")
    return {
        "employees_count": len(rows),
        "totals": {
            "employees_count": len(rows),
            "operations_count": len(rows),
            "total_amount": sum(r.get("Net à payer", 0) for r in rows),
        },
        "anomalies": [],
        "warnings": warnings,
        "can_generate": True,
        "details": {"lines": rows},
    }


def generate_attestations_export(
    company_id: str,
    period: str,
    employee_ids: Optional[List[str]] = None,
    file_format: str = "xlsx",
) -> bytes:
    rows = _fetch_attestation_rows(company_id, period, employee_ids)
    if file_format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Attestations"
        for col_idx, h in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, h in enumerate(HEADERS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    return generate_csv(rows, HEADERS)
