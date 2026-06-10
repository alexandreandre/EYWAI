# Export « Acomptes & avances » — liste détaillée + écritures comptables OD (425x).
from __future__ import annotations

import io
from calendar import monthrange
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.database import supabase
from app.modules.saisies_avances.infrastructure.queries import (
    list_advance_payments_by_period,
    list_advance_repayments_by_period,
)
from app.shared.utils.export import format_period, generate_csv

LIST_HEADERS = [
    "Employé",
    "Nature",
    "Compte comptable",
    "Évènement",
    "Montant versé",
    "Montant remboursé",
    "Date",
    "Statut",
    "Prime",
]

ECRITURES_HEADERS = [
    "Date écriture",
    "Journal",
    "Compte comptable",
    "Libellé",
    "Débit",
    "Crédit",
    "Référence export",
    "Période de paie",
]

DEFAULT_BANK_ACCOUNT = "512000"
DEFAULT_NET_ACCOUNT = "425000"


def _round2(value: float) -> float:
    return round(value, 2)


def get_bank_account(company_id: str) -> str:
    """Compte banque pour les versements d'acomptes (config ou défaut)."""
    config_data: dict | None = None
    if company_id:
        r = (
            supabase.table("payroll_config")
            .select("config_data")
            .eq("config_key", "comptes_avances_acomptes")
            .eq("company_id", company_id)
            .eq("is_active", True)
            .limit(1)
            .execute()
        )
        if r.data:
            config_data = r.data[0].get("config_data") or {}
    if not config_data:
        r = (
            supabase.table("payroll_config")
            .select("config_data")
            .eq("config_key", "comptes_avances_acomptes")
            .is_("company_id", "null")
            .eq("is_active", True)
            .limit(1)
            .execute()
        )
        if r.data:
            config_data = r.data[0].get("config_data") or {}
    if config_data:
        for key in ("banque", "512", "compte_banque"):
            if key in config_data:
                return str(config_data[key])
    return DEFAULT_BANK_ACCOUNT


def _period_end_date(period: str) -> str:
    year, month = map(int, period.split("-"))
    last_day = monthrange(year, month)[1]
    return f"{year}-{month:02d}-{last_day:02d}"


def _status_label(status: Optional[str]) -> str:
    labels = {
        "pending": "En attente",
        "approved": "À verser",
        "rejected": "Rejetée",
        "paid": "Versée",
    }
    return labels.get(status or "", status or "")


def _build_list_rows(
    payments: List[Dict[str, Any]],
    repayments: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    for item in payments:
        rows.append(
            {
                "Employé": item.get("employee_name", ""),
                "Nature": item.get("advance_type_label", ""),
                "Compte comptable": item.get("accounting_account", ""),
                "Évènement": "Versement",
                "Montant versé": _round2(float(item.get("amount_paid", 0) or 0)),
                "Montant remboursé": 0.0,
                "Date": item.get("event_date", ""),
                "Statut": _status_label(item.get("advance_status")),
                "Prime": item.get("prime_label") or "",
            }
        )

    for item in repayments:
        rows.append(
            {
                "Employé": item.get("employee_name", ""),
                "Nature": item.get("advance_type_label", ""),
                "Compte comptable": item.get("accounting_account", ""),
                "Évènement": "Remboursement",
                "Montant versé": 0.0,
                "Montant remboursé": _round2(float(item.get("amount_repaid", 0) or 0)),
                "Date": item.get("event_date", ""),
                "Statut": _status_label(item.get("advance_status")),
                "Prime": item.get("prime_label") or "",
            }
        )

    return rows


def get_acomptes_data(
    company_id: str,
    period: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Retourne versements, remboursements, lignes liste et totaux."""
    payments = list_advance_payments_by_period(company_id, period)
    repayments = list_advance_repayments_by_period(company_id, period)
    list_rows = _build_list_rows(payments, repayments)

    total_paid = sum(float(p.get("amount_paid", 0) or 0) for p in payments)
    total_repaid = sum(float(r.get("amount_repaid", 0) or 0) for r in repayments)

    totals_by_account: Dict[str, Dict[str, float]] = {}
    for item in payments + repayments:
        compte = str(item.get("accounting_account") or "425")
        if compte not in totals_by_account:
            totals_by_account[compte] = {"versements": 0.0, "remboursements": 0.0}
        if item.get("event_type") == "versement":
            totals_by_account[compte]["versements"] += float(
                item.get("amount_paid", 0) or 0
            )
        else:
            totals_by_account[compte]["remboursements"] += float(
                item.get("amount_repaid", 0) or 0
            )

    employee_ids = {
        str(p.get("employee_id"))
        for p in payments + repayments
        if p.get("employee_id")
    }

    totals = {
        "employees_count": len(employee_ids),
        "total_amount": _round2(total_paid + total_repaid),
        "total_versements": _round2(total_paid),
        "total_remboursements": _round2(total_repaid),
        "operations_count": len(payments) + len(repayments),
        "totals_by_account": totals_by_account,
    }
    return payments, repayments, list_rows, totals


def generate_acomptes_ecritures(
    company_id: str,
    period: str,
    payments: Optional[List[Dict[str, Any]]] = None,
    repayments: Optional[List[Dict[str, Any]]] = None,
    date_ecriture: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Génère les écritures OD pour versements et remboursements d'acomptes."""
    if payments is None or repayments is None:
        payments, repayments, _, _ = get_acomptes_data(company_id, period)

    if not date_ecriture:
        date_ecriture = _period_end_date(period)

    bank_account: Optional[str] = None
    period_label = format_period(period)
    ecritures: List[Dict[str, Any]] = []

    for payment in payments:
        if bank_account is None:
            bank_account = get_bank_account(company_id)
        compte = str(payment.get("accounting_account") or "425")
        montant = float(payment.get("amount_paid", 0) or 0)
        if montant <= 0:
            continue
        employee = payment.get("employee_name", "")
        nature = payment.get("advance_type_label", "Acompte")
        libelle = f"{nature} — {employee} — {period_label}"
        ecritures.append(
            {
                "date_ecriture": payment.get("event_date") or date_ecriture,
                "journal": "OD",
                "compte_comptable": compte,
                "libelle": libelle,
                "debit": _round2(montant),
                "credit": 0.0,
                "reference_export": f"ACOMPTE_V_{period}",
                "periode_paie": period,
            }
        )
        ecritures.append(
            {
                "date_ecriture": payment.get("event_date") or date_ecriture,
                "journal": "OD",
                "compte_comptable": bank_account or DEFAULT_BANK_ACCOUNT,
                "libelle": f"Virement {nature} — {employee}",
                "debit": 0.0,
                "credit": _round2(montant),
                "reference_export": f"ACOMPTE_V_{period}",
                "periode_paie": period,
            }
        )

    for repayment in repayments:
        compte = str(repayment.get("accounting_account") or "425")
        montant = float(repayment.get("amount_repaid", 0) or 0)
        if montant <= 0:
            continue
        employee = repayment.get("employee_name", "")
        nature = repayment.get("advance_type_label", "Acompte")
        libelle = f"Remboursement {nature} — {employee} — {period_label}"
        ecritures.append(
            {
                "date_ecriture": date_ecriture,
                "journal": "OD",
                "compte_comptable": DEFAULT_NET_ACCOUNT,
                "libelle": libelle,
                "debit": _round2(montant),
                "credit": 0.0,
                "reference_export": f"ACOMPTE_R_{period}",
                "periode_paie": period,
            }
        )
        ecritures.append(
            {
                "date_ecriture": date_ecriture,
                "journal": "OD",
                "compte_comptable": compte,
                "libelle": libelle,
                "debit": 0.0,
                "credit": _round2(montant),
                "reference_export": f"ACOMPTE_R_{period}",
                "periode_paie": period,
            }
        )

    return ecritures


def get_repayments_total_by_account(
    company_id: str, period: str
) -> Dict[str, float]:
    """Totaux remboursements par compte 425x pour intégration OD globale."""
    _, repayments, _, _ = get_acomptes_data(company_id, period)
    totals: Dict[str, float] = {}
    for rep in repayments:
        compte = str(rep.get("accounting_account") or "425")
        amt = float(rep.get("amount_repaid", 0) or 0)
        if amt > 0:
            totals[compte] = totals.get(compte, 0.0) + amt
    return totals


def _write_sheet(ws, headers: List[str], rows: List[Dict[str, Any]], title: str) -> None:
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=12)

    ws.cell(row=1, column=1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(rows, start=4):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

    for col_idx, header in enumerate(headers, start=1):
        max_length = max(
            len(str(header)),
            max((len(str(row.get(header, ""))) for row in rows), default=0),
        )
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = min(
            max_length + 2, 50
        )


def _ecritures_to_export_rows(ecritures: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "Date écriture": e["date_ecriture"],
            "Journal": e["journal"],
            "Compte comptable": e["compte_comptable"],
            "Libellé": e["libelle"],
            "Débit": e["debit"],
            "Crédit": e["credit"],
            "Référence export": e.get("reference_export", ""),
            "Période de paie": e["periode_paie"],
        }
        for e in ecritures
    ]


def _generate_xlsx_workbook(
    list_rows: List[Dict[str, Any]],
    ecritures: List[Dict[str, Any]],
    period: str,
) -> bytes:
    wb = Workbook()
    period_label = format_period(period)

    ws_list = wb.active
    ws_list.title = "Liste acomptes"
    _write_sheet(
        ws_list,
        LIST_HEADERS,
        list_rows,
        f"Acomptes & avances — {period_label}",
    )

    ws_od = wb.create_sheet("Écritures comptables")
    _write_sheet(
        ws_od,
        ECRITURES_HEADERS,
        _ecritures_to_export_rows(ecritures),
        f"Écritures OD acomptes — {period_label}",
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def preview_acomptes(company_id: str, period: str) -> Dict[str, Any]:
    payments, repayments, list_rows, totals = get_acomptes_data(company_id, period)
    ecritures = generate_acomptes_ecritures(
        company_id, period, payments, repayments
    )

    anomalies: List[Dict[str, Any]] = []
    warnings: List[str] = []

    if totals["operations_count"] == 0:
        warnings.append(
            "Aucun versement ni remboursement d'acompte sur cette période."
        )

    for item in payments + repayments:
        if not item.get("accounting_account"):
            anomalies.append(
                {
                    "type": "warning",
                    "message": (
                        f"Compte comptable manquant pour {item.get('employee_name', 'employé')} "
                        f"({item.get('advance_type_label', 'acompte')})"
                    ),
                    "severity": "warning",
                    "employee_id": item.get("employee_id"),
                    "employee_name": item.get("employee_name"),
                }
            )

    missing_account = [
        a for a in anomalies if a.get("severity") == "blocking"
    ]
    can_generate = len(missing_account) == 0

    return {
        "employees_count": totals["employees_count"],
        "totals": totals,
        "anomalies": anomalies,
        "warnings": warnings,
        "can_generate": can_generate,
        "details": {
            "lines": list_rows,
            "ecritures_count": len(ecritures),
        },
    }


def generate_acomptes_export(
    company_id: str,
    period: str,
    file_format: str = "csv",
) -> bytes:
    """Génère le fichier export (xlsx multi-feuilles ou csv liste)."""
    payments, repayments, list_rows, _ = get_acomptes_data(company_id, period)
    ecritures = generate_acomptes_ecritures(
        company_id, period, payments, repayments
    )

    if file_format == "xlsx":
        return _generate_xlsx_workbook(list_rows, ecritures, period)
    return generate_csv(list_rows, LIST_HEADERS)


def generate_acomptes_ecritures_export(
    company_id: str,
    period: str,
    file_format: str = "csv",
) -> bytes:
    """Génère uniquement le fichier des écritures comptables."""
    ecritures = generate_acomptes_ecritures(company_id, period)
    rows = _ecritures_to_export_rows(ecritures)
    if file_format == "xlsx":
        from app.shared.utils.export import generate_xlsx

        return generate_xlsx(
            rows,
            ECRITURES_HEADERS,
            f"Écritures acomptes {format_period(period)}",
        )
    return generate_csv(rows, ECRITURES_HEADERS)
