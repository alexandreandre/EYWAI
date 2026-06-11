# Export « Saisies sur salaire » — liste des prélèvements + écritures comptables OD (427x).
from __future__ import annotations

import io
from calendar import monthrange
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.modules.saisies_avances.infrastructure.queries import (
    list_seizure_deductions_by_period,
)
from app.shared.utils.export import format_period, generate_csv

LIST_HEADERS = [
    "Employé",
    "Type de saisie",
    "Créancier",
    "Compte comptable",
    "Montant prélevé",
    "Quotité saisissable",
    "Net avant saisie",
    "Période",
    "Statut saisie",
    "Référence",
]

ECRITURES_HEADERS = [
    "Date écriture",
    "Journal",
    "Compte comptable",
    "Libellé",
    "Débit",
    "Crédit",
    "Référence export",
    "Période de paie",
]

DEFAULT_NET_ACCOUNT = "425000"


def _round2(value: float) -> float:
    return round(value, 2)


def _period_end_date(period: str) -> str:
    year, month = map(int, period.split("-"))
    last_day = monthrange(year, month)[1]
    return f"{year}-{month:02d}-{last_day:02d}"


def _status_label(status: Optional[str]) -> str:
    labels = {
        "active": "Active",
        "suspended": "Suspendue",
        "closed": "Clôturée",
    }
    return labels.get(status or "", status or "")


def _build_list_rows(deductions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for item in deductions:
        rows.append(
            {
                "Employé": item.get("employee_name", ""),
                "Type de saisie": item.get("seizure_type_label", ""),
                "Créancier": item.get("creditor_name", ""),
                "Compte comptable": item.get("accounting_account", ""),
                "Montant prélevé": _round2(float(item.get("deducted_amount", 0) or 0)),
                "Quotité saisissable": _round2(
                    float(item.get("seizable_amount", 0) or 0)
                ),
                "Net avant saisie": _round2(float(item.get("net_salary", 0) or 0)),
                "Période": item.get("period", ""),
                "Statut saisie": _status_label(item.get("seizure_status")),
                "Référence": f"SAISIE_{item.get('period', '')}_{item.get('seizure_id', '')[:8]}",
            }
        )
    return rows


def get_saisies_data(
    company_id: str,
    period: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Retourne prélèvements, lignes liste et totaux."""
    deductions = list_seizure_deductions_by_period(company_id, period)
    list_rows = _build_list_rows(deductions)

    total_deducted = sum(float(d.get("deducted_amount", 0) or 0) for d in deductions)
    employee_ids = {
        str(d.get("employee_id")) for d in deductions if d.get("employee_id")
    }

    totals_by_account: Dict[str, Dict[str, float]] = {}
    for item in deductions:
        compte = str(item.get("accounting_account") or "427000")
        if compte not in totals_by_account:
            totals_by_account[compte] = {"prelevements": 0.0}
        totals_by_account[compte]["prelevements"] += float(
            item.get("deducted_amount", 0) or 0
        )

    totals = {
        "employees_count": len(employee_ids),
        "total_amount": _round2(total_deducted),
        "total_prelevements": _round2(total_deducted),
        "operations_count": len(deductions),
        "totals_by_account": totals_by_account,
    }
    return deductions, list_rows, totals


def generate_saisies_ecritures(
    company_id: str,
    period: str,
    deductions: Optional[List[Dict[str, Any]]] = None,
    date_ecriture: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Génère les écritures OD pour les prélèvements de saisies sur bulletin.

    Schéma PCG classique :
    - Débit 425 (net à payer) — retenue sur salaire
    - Crédit 427x (Personnel — Oppositions) — dette envers le créancier
    """
    if deductions is None:
        deductions, _, _ = get_saisies_data(company_id, period)

    if not date_ecriture:
        date_ecriture = _period_end_date(period)

    period_label = format_period(period)
    ecritures: List[Dict[str, Any]] = []

    for deduction in deductions:
        montant = float(deduction.get("deducted_amount", 0) or 0)
        if montant <= 0:
            continue

        employee = deduction.get("employee_name", "")
        nature = deduction.get("seizure_type_label", "Saisie")
        creditor = deduction.get("creditor_name", "")
        compte = str(deduction.get("accounting_account") or "427000")
        libelle_parts = [nature, employee]
        if creditor:
            libelle_parts.append(creditor)
        libelle = " — ".join(p for p in libelle_parts if p) + f" — {period_label}"
        reference = f"SAISIE_{period}"

        ecritures.append(
            {
                "date_ecriture": date_ecriture,
                "journal": "OD",
                "compte_comptable": DEFAULT_NET_ACCOUNT,
                "libelle": libelle,
                "debit": _round2(montant),
                "credit": 0.0,
                "reference_export": reference,
                "periode_paie": period,
            }
        )
        ecritures.append(
            {
                "date_ecriture": date_ecriture,
                "journal": "OD",
                "compte_comptable": compte,
                "libelle": libelle,
                "debit": 0.0,
                "credit": _round2(montant),
                "reference_export": reference,
                "periode_paie": period,
            }
        )

    return ecritures


def get_seizures_total_by_account(company_id: str, period: str) -> Dict[str, float]:
    """Totaux prélèvements par compte 427x pour intégration OD globale."""
    deductions, _, _ = get_saisies_data(company_id, period)
    totals: Dict[str, float] = {}
    for item in deductions:
        compte = str(item.get("accounting_account") or "427000")
        amt = float(item.get("deducted_amount", 0) or 0)
        if amt > 0:
            totals[compte] = totals.get(compte, 0.0) + amt
    return totals


def _write_sheet(ws, headers: List[str], rows: List[Dict[str, Any]], title: str) -> None:
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=12)

    ws.cell(row=1, column=1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(rows, start=4):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

    for col_idx, header in enumerate(headers, start=1):
        max_length = max(
            len(str(header)),
            max((len(str(row.get(header, ""))) for row in rows), default=0),
        )
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = min(
            max_length + 2, 50
        )


def _ecritures_to_export_rows(ecritures: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "Date écriture": e["date_ecriture"],
            "Journal": e["journal"],
            "Compte comptable": e["compte_comptable"],
            "Libellé": e["libelle"],
            "Débit": e["debit"],
            "Crédit": e["credit"],
            "Référence export": e.get("reference_export", ""),
            "Période de paie": e["periode_paie"],
        }
        for e in ecritures
    ]


def _generate_xlsx_workbook(
    list_rows: List[Dict[str, Any]],
    ecritures: List[Dict[str, Any]],
    period: str,
) -> bytes:
    wb = Workbook()
    period_label = format_period(period)

    ws_list = wb.active
    ws_list.title = "Liste saisies"
    _write_sheet(
        ws_list,
        LIST_HEADERS,
        list_rows,
        f"Saisies sur salaire — {period_label}",
    )

    ws_od = wb.create_sheet("Écritures comptables")
    _write_sheet(
        ws_od,
        ECRITURES_HEADERS,
        _ecritures_to_export_rows(ecritures),
        f"Écritures OD saisies — {period_label}",
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def preview_saisies(company_id: str, period: str) -> Dict[str, Any]:
    deductions, list_rows, totals = get_saisies_data(company_id, period)
    ecritures = generate_saisies_ecritures(company_id, period, deductions)

    anomalies: List[Dict[str, Any]] = []
    warnings: List[str] = []

    if totals["operations_count"] == 0:
        warnings.append(
            "Aucun prélèvement de saisie sur salaire enregistré sur cette période."
        )

    for item in deductions:
        if not item.get("creditor_name"):
            anomalies.append(
                {
                    "type": "warning",
                    "message": (
                        f"Créancier manquant pour {item.get('employee_name', 'employé')} "
                        f"({item.get('seizure_type_label', 'saisie')})"
                    ),
                    "severity": "warning",
                    "employee_id": item.get("employee_id"),
                    "employee_name": item.get("employee_name"),
                }
            )

    can_generate = True

    return {
        "employees_count": totals["employees_count"],
        "totals": totals,
        "anomalies": anomalies,
        "warnings": warnings,
        "can_generate": can_generate,
        "details": {
            "lines": list_rows,
            "ecritures_count": len(ecritures),
        },
    }


def generate_saisies_export(
    company_id: str,
    period: str,
    file_format: str = "csv",
) -> bytes:
    """Génère le fichier export (xlsx multi-feuilles ou csv liste)."""
    deductions, list_rows, _ = get_saisies_data(company_id, period)
    ecritures = generate_saisies_ecritures(company_id, period, deductions)

    if file_format == "xlsx":
        return _generate_xlsx_workbook(list_rows, ecritures, period)
    return generate_csv(list_rows, LIST_HEADERS)


def generate_saisies_ecritures_export(
    company_id: str,
    period: str,
    file_format: str = "csv",
) -> bytes:
    """Génère uniquement le fichier des écritures comptables."""
    ecritures = generate_saisies_ecritures(company_id, period)
    rows = _ecritures_to_export_rows(ecritures)
    if file_format == "xlsx":
        from app.shared.utils.export import generate_xlsx

        return generate_xlsx(
            rows,
            ECRITURES_HEADERS,
            f"Écritures saisies {format_period(period)}",
        )
    return generate_csv(rows, ECRITURES_HEADERS)
