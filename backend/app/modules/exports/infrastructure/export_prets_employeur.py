# Export « Prêts employeur » — remboursements sur bulletin + écritures 274x.
from __future__ import annotations

import io
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font

from app.modules.exports.infrastructure.payroll_ledger import (
    build_payroll_ledger,
    ledger_to_od_export_rows,
)
from app.shared.utils.export import generate_csv

LIST_HEADERS = [
    "Employé",
    "Capital remboursé",
    "Intérêts",
    "Total prélevé",
    "Période",
]

ECRITURES_HEADERS = [
    "Date écriture",
    "Journal",
    "Compte comptable",
    "Libellé",
    "Débit",
    "Crédit",
    "Référence export",
    "Période de paie",
]


def _list_loan_repayments(company_id: str, period: str) -> List[Dict[str, Any]]:
    from app.modules.exports.infrastructure.payroll_ledger import (
        list_loan_repayments_by_period,
    )
    return list_loan_repayments_by_period(company_id, period)


def get_prets_data(
    company_id: str, period: str
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    repayments = _list_loan_repayments(company_id, period)
    list_rows = [
        {
            "Employé": r.get("employee_name", ""),
            "Capital remboursé": round(float(r.get("capital_amount", 0) or 0), 2),
            "Intérêts": round(float(r.get("interest_amount", 0) or 0), 2),
            "Total prélevé": round(float(r.get("total_amount", 0) or 0), 2),
            "Période": period,
        }
        for r in repayments
    ]
    total = sum(float(r.get("total_amount", 0) or 0) for r in repayments)
    employee_ids = {str(r.get("employee_id")) for r in repayments if r.get("employee_id")}
    totals = {
        "employees_count": len(employee_ids),
        "total_amount": round(total, 2),
        "total_prelevements": round(total, 2),
        "operations_count": len(repayments),
    }
    return repayments, list_rows, totals


def generate_prets_ecritures(
    company_id: str, period: str
) -> List[Dict[str, Any]]:
    ecritures_raw, _, _ = build_payroll_ledger(
        company_id, period, scope="auxiliaries"
    )
    ecritures = ledger_to_od_export_rows(ecritures_raw)
    return [e for e in ecritures if "prêt" in e.get("libelle", "").lower()]


def preview_prets_employeur(company_id: str, period: str) -> Dict[str, Any]:
    _, list_rows, totals = get_prets_data(company_id, period)
    ecritures = generate_prets_ecritures(company_id, period)
    warnings: List[str] = []
    if totals["operations_count"] == 0:
        warnings.append("Aucun remboursement de prêt employeur sur cette période.")
    return {
        "employees_count": totals["employees_count"],
        "totals": totals,
        "anomalies": [],
        "warnings": warnings,
        "can_generate": True,
        "details": {"lines": list_rows, "ecritures_count": len(ecritures)},
    }


def generate_prets_employeur_export(
    company_id: str, period: str, file_format: str = "xlsx"
) -> bytes:
    _, list_rows, _ = get_prets_data(company_id, period)
    ecritures = generate_prets_ecritures(company_id, period)
    if file_format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Liste prêts"
        for col_idx, h in enumerate(LIST_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h).font = Font(bold=True)
        for row_idx, row in enumerate(list_rows, start=2):
            for col_idx, h in enumerate(LIST_HEADERS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
        ws2 = wb.create_sheet("Écritures comptables")
        for col_idx, h in enumerate(ECRITURES_HEADERS, start=1):
            ws2.cell(row=1, column=col_idx, value=h).font = Font(bold=True)
        for row_idx, e in enumerate(ecritures, start=2):
            ws2.cell(row=row_idx, column=1, value=e["date_ecriture"])
            ws2.cell(row=row_idx, column=2, value=e["journal"])
            ws2.cell(row=row_idx, column=3, value=e["compte_comptable"])
            ws2.cell(row=row_idx, column=4, value=e["libelle"])
            ws2.cell(row=row_idx, column=5, value=e["debit"])
            ws2.cell(row=row_idx, column=6, value=e["credit"])
            ws2.cell(row=row_idx, column=7, value=e.get("reference_export", ""))
            ws2.cell(row=row_idx, column=8, value=e["periode_paie"])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    return generate_csv(list_rows, LIST_HEADERS)
