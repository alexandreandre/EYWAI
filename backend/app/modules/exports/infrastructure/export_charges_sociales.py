# Export « Charges sociales par caisse » — tableau détaillé par organisme.
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.modules.exports.domain.charges_organisme import resolve_organisme
from app.modules.exports.infrastructure.export_ecritures_comptables import (
    get_payslip_data_for_od,
)
from app.shared.utils.export import format_period, generate_csv

DETAIL_HEADERS = [
    "Organisme",
    "Libellé cotisation",
    "Part salariale",
    "Part patronale",
    "Total cotisations",
]

SUMMARY_HEADERS = [
    "Organisme",
    "Nombre de salariés",
    "Part salariale",
    "Part patronale",
    "Total cotisations",
]


def _round2(value: float) -> float:
    return round(value, 2)


def _aggregate_charges(
    payslip_list: List[Dict[str, Any]],
    caisses: Optional[List[str]] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Agrège les cotisations par ligne de détail et par organisme."""
    caisses_filter = {c.upper() for c in caisses} if caisses else None

    detail_map: Dict[str, Dict[str, Any]] = {}
    summary_map: Dict[str, Dict[str, Any]] = {}

    totals = {
        "employees_count": len(payslip_list),
        "total_cotisations_salariales": 0.0,
        "total_cotisations_patronales": 0.0,
    }

    for payslip in payslip_list:
        employee_id = payslip.get("employee_id")
        for coti in payslip.get("cotisations_detail", []):
            if not isinstance(coti, dict):
                continue

            libelle = coti.get("libelle", "Cotisation")
            organisme = resolve_organisme(libelle)
            if caisses_filter and organisme not in caisses_filter:
                continue

            montant_salarial = float(coti.get("montant_salarial", 0) or 0)
            montant_patronal = float(coti.get("montant_patronal", 0) or 0)
            if montant_salarial == 0 and montant_patronal == 0:
                continue

            detail_key = f"{organisme}::{libelle}"
            if detail_key not in detail_map:
                detail_map[detail_key] = {
                    "Organisme": organisme,
                    "Libellé cotisation": libelle,
                    "Part salariale": 0.0,
                    "Part patronale": 0.0,
                    "Total cotisations": 0.0,
                }
            detail_map[detail_key]["Part salariale"] += montant_salarial
            detail_map[detail_key]["Part patronale"] += montant_patronal
            detail_map[detail_key]["Total cotisations"] += montant_salarial + montant_patronal

            if organisme not in summary_map:
                summary_map[organisme] = {
                    "Organisme": organisme,
                    "Nombre de salariés": set(),
                    "Part salariale": 0.0,
                    "Part patronale": 0.0,
                    "Total cotisations": 0.0,
                }
            if employee_id:
                summary_map[organisme]["Nombre de salariés"].add(employee_id)
            summary_map[organisme]["Part salariale"] += montant_salarial
            summary_map[organisme]["Part patronale"] += montant_patronal
            summary_map[organisme]["Total cotisations"] += montant_salarial + montant_patronal

            totals["total_cotisations_salariales"] += montant_salarial
            totals["total_cotisations_patronales"] += montant_patronal

    detail_rows: List[Dict[str, Any]] = []
    for row in sorted(detail_map.values(), key=lambda r: (r["Organisme"], r["Libellé cotisation"])):
        detail_rows.append(
            {
                "Organisme": row["Organisme"],
                "Libellé cotisation": row["Libellé cotisation"],
                "Part salariale": _round2(row["Part salariale"]),
                "Part patronale": _round2(row["Part patronale"]),
                "Total cotisations": _round2(row["Total cotisations"]),
            }
        )

    summary_rows: List[Dict[str, Any]] = []
    for organisme in sorted(summary_map.keys()):
        row = summary_map[organisme]
        salaries: Set[Any] = row["Nombre de salariés"]
        summary_rows.append(
            {
                "Organisme": organisme,
                "Nombre de salariés": len(salaries),
                "Part salariale": _round2(row["Part salariale"]),
                "Part patronale": _round2(row["Part patronale"]),
                "Total cotisations": _round2(row["Total cotisations"]),
            }
        )

    totals["total_cotisations_salariales"] = _round2(totals["total_cotisations_salariales"])
    totals["total_cotisations_patronales"] = _round2(totals["total_cotisations_patronales"])
    totals["total_amount"] = _round2(
        totals["total_cotisations_salariales"] + totals["total_cotisations_patronales"]
    )

    return detail_rows, summary_rows, totals


def get_charges_sociales_data(
    company_id: str,
    period: str,
    employee_ids: Optional[List[str]] = None,
    caisses: Optional[List[str]] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    payslip_list, _ = get_payslip_data_for_od(
        company_id, period, employee_ids, "charges_sociales"
    )
    return _aggregate_charges(payslip_list, caisses)


def _write_sheet(
    ws,
    headers: List[str],
    rows: List[Dict[str, Any]],
    title: str,
) -> None:
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=12)

    ws.cell(row=1, column=1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(rows, start=4):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

    for col_idx, header in enumerate(headers, start=1):
        max_length = max(
            len(str(header)),
            max((len(str(row.get(header, ""))) for row in rows), default=0),
        )
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = min(
            max_length + 2, 50
        )


def _generate_xlsx_workbook(
    detail_rows: List[Dict[str, Any]],
    summary_rows: List[Dict[str, Any]],
    period: str,
    include_consolidated: bool,
) -> bytes:
    wb = Workbook()
    period_label = format_period(period)

    if include_consolidated and summary_rows:
        ws_summary = wb.active
        ws_summary.title = "Synthèse par organisme"
        _write_sheet(
            ws_summary,
            SUMMARY_HEADERS,
            summary_rows,
            f"Synthèse par organisme — {period_label}",
        )
        ws_detail = wb.create_sheet("Détail par cotisation")
    else:
        ws_detail = wb.active
        ws_detail.title = "Détail par cotisation"

    _write_sheet(
        ws_detail,
        DETAIL_HEADERS,
        detail_rows,
        f"Détail par cotisation — {period_label}",
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def preview_charges_sociales(
    company_id: str,
    period: str,
    employee_ids: Optional[List[str]] = None,
    caisses: Optional[List[str]] = None,
    include_consolidated: bool = True,
) -> Dict[str, Any]:
    detail_rows, summary_rows, totals = get_charges_sociales_data(
        company_id, period, employee_ids, caisses
    )

    anomalies: List[Dict[str, Any]] = []
    warnings: List[str] = []

    if totals["employees_count"] == 0:
        warnings.append("Aucun bulletin trouvé pour cette période")
    if totals["employees_count"] > 0 and not detail_rows:
        warnings.append("Aucune cotisation trouvée sur les bulletins de la période")

    organismes = [
        {
            "organisme": row["Organisme"],
            "nombre_salaries": row["Nombre de salariés"],
            "total_cotisations_salariales": row["Part salariale"],
            "total_cotisations_patronales": row["Part patronale"],
            "total_cotisations": row["Total cotisations"],
        }
        for row in summary_rows
    ]

    return {
        "employees_count": totals["employees_count"],
        "totals": totals,
        "anomalies": anomalies,
        "warnings": warnings,
        "can_generate": len([a for a in anomalies if a.get("severity") == "blocking"]) == 0,
        "details": {
            "organismes": organismes,
            "lines": detail_rows,
            "include_consolidated": include_consolidated,
        },
    }


def generate_charges_sociales_export(
    company_id: str,
    period: str,
    employee_ids: Optional[List[str]] = None,
    format: str = "xlsx",
    caisses: Optional[List[str]] = None,
    include_consolidated: bool = True,
) -> bytes:
    detail_rows, summary_rows, _ = get_charges_sociales_data(
        company_id, period, employee_ids, caisses
    )

    if format == "xlsx":
        return _generate_xlsx_workbook(
            detail_rows, summary_rows, period, include_consolidated
        )
    return generate_csv(detail_rows, DETAIL_HEADERS)
