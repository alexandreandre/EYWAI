# app/modules/cse/infrastructure/cse_export_impl.py
"""
Exports Excel CSE (base élus, heures délégation, historique réunions).
Implémentation autonome ex-services.cse_export_service.
"""

from __future__ import annotations

import io
from datetime import date, datetime, time
from typing import Any, Dict, List, Optional

from app.shared.utils.export import generate_xlsx


def _vers_date(valeur: Any) -> Optional[date]:
    """Normalise une valeur date/datetime/chaîne ISO vers une date. None si illisible.

    Les exports reçoivent des objets date (model_dump) ou des chaînes ISO (dicts bruts) :
    il faut accepter les deux plutôt que d'avaler une TypeError.
    """
    if valeur is None or valeur == "":
        return None
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    if isinstance(valeur, str):
        try:
            return datetime.fromisoformat(valeur).date()
        except ValueError:
            return None
    return None


def _formater_date(valeur: Any) -> str:
    """Rend une date au format JJ/MM/AAAA, ou la valeur telle quelle si illisible."""
    jour = _vers_date(valeur)
    if jour is None:
        return "" if valeur is None else str(valeur)
    return jour.strftime("%d/%m/%Y")


def _formater_heure(valeur: Any) -> str:
    """Rend une heure au format HH:MM, "" si absente.

    Les réunions reçoivent un `datetime.time` (model_dump en mode python) ou une chaîne
    "HH:MM:SS"/"HH:MM" (dict brut) : il faut accepter les deux plutôt que de trancher un
    `time` avec `[:5]`, ce qui lève TypeError.
    """
    if valeur is None or valeur == "":
        return ""
    if isinstance(valeur, datetime):
        return valeur.strftime("%H:%M")
    if isinstance(valeur, time):
        return valeur.strftime("%H:%M")
    if isinstance(valeur, str):
        texte = valeur.strip()
        for motif in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(texte, motif).strftime("%H:%M")
            except ValueError:
                continue
        return texte
    return str(valeur)


def export_elected_members(members: List[Dict[str, Any]]) -> bytes:
    """Export Excel de la base des élus CSE."""
    headers = [
        "Nom",
        "Prénom",
        "Poste",
        "Rôle CSE",
        "Collège",
        "Date début mandat",
        "Date fin mandat",
        "Jours restants",
        "Statut",
    ]
    aujourd_hui = date.today()
    data = []
    for member in members:
        fin = _vers_date(member.get("end_date"))
        days_remaining = (fin - aujourd_hui).days if fin is not None else None
        # La date de fin prime toujours : un mandat arrivé à son terme est « Expiré »,
        # jamais « Révoqué », quelle que soit la valeur de is_active (le script d'import
        # pose is_active=False précisément parce que la date de fin est passée).
        # is_active explicitement à faux ne compte que pour un mandat interrompu avant
        # son terme. La clé peut être absente (dicts partiels en test) : .get(...) renvoie
        # alors None, pas False, et on retombe sur la logique par date.
        if days_remaining is not None and days_remaining < 0:
            status = "Expiré"
        elif member.get("is_active") is False:
            status = "Révoqué"
        elif days_remaining is None:
            status = "Inconnu"
        elif days_remaining <= 90:
            status = "Expire bientôt"
        else:
            status = "Actif"
        data.append(
            {
                "Nom": member.get("last_name", ""),
                "Prénom": member.get("first_name", ""),
                "Poste": member.get("job_title", ""),
                "Rôle CSE": (member.get("role") or "").capitalize(),
                "Collège": member.get("college", ""),
                "Date début mandat": _formater_date(member.get("start_date")),
                "Date fin mandat": _formater_date(member.get("end_date")),
                "Jours restants": days_remaining if days_remaining is not None else "",
                "Statut": status,
            }
        )
    return generate_xlsx(data, headers, "Base élus CSE")


def export_delegation_hours(
    hours: List[Dict[str, Any]], summary: List[Dict[str, Any]]
) -> bytes:
    """Export Excel des heures de délégation."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Récapitulatif"
    headers_summary = [
        "Élu",
        "Rôle",
        "Crédit base (h)",
        "Reporté dispo (h)",
        "Mutualisé reçu (h)",
        "Plafond mensuel (h)",
        "Disponible (h)",
        "Consommé (h)",
        "Restant (h)",
        "Dépassement (h)",
        "Période",
    ]
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(headers_summary, start=1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, item in enumerate(summary, start=2):
        ws_summary.cell(
            row=row_idx,
            column=1,
            value=f"{item.get('first_name', '')} {item.get('last_name', '')}",
        )
        ws_summary.cell(
            row=row_idx, column=2, value=item.get("role", "")
        )
        ws_summary.cell(
            row=row_idx, column=3, value=item.get("credit_base", item.get("quota_hours_per_month", 0))
        )
        ws_summary.cell(row=row_idx, column=4, value=item.get("reported_available", 0))
        ws_summary.cell(row=row_idx, column=5, value=item.get("transfers_in", 0))
        ws_summary.cell(row=row_idx, column=6, value=item.get("monthly_cap", 0))
        ws_summary.cell(row=row_idx, column=7, value=item.get("available_hours", 0))
        ws_summary.cell(row=row_idx, column=8, value=item.get("consumed_hours", 0))
        ws_summary.cell(row=row_idx, column=9, value=item.get("remaining_hours", 0))
        ws_summary.cell(row=row_idx, column=10, value=item.get("overrun_hours", 0))
        period_start = item.get("period_start")
        period_end = item.get("period_end")
        if period_start and period_end:
            ws_summary.cell(
                row=row_idx,
                column=11,
                value=f"{_formater_date(period_start)} - {_formater_date(period_end)}",
            )
    for col_idx in range(1, len(headers_summary) + 1):
        h = headers_summary[col_idx - 1]
        max_length = max(
            len(str(h)),
            max(
                (
                    len(str(ws_summary.cell(row=row_idx, column=col_idx).value or ""))
                    for row_idx in range(2, ws_summary.max_row + 1)
                ),
                default=0,
            ),
        )
        ws_summary.column_dimensions[
            ws_summary.cell(row=1, column=col_idx).column_letter
        ].width = min(max_length + 2, 50)
    ws_detail = wb.create_sheet("Détail heures")
    headers_detail = ["Date", "Élu", "Durée (h)", "Source", "Motif", "Réunion associée"]
    for col_idx, header in enumerate(headers_detail, start=1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, hour in enumerate(hours, start=2):
        ws_detail.cell(row=row_idx, column=1, value=_formater_date(hour.get("date")))
        ws_detail.cell(
            row=row_idx,
            column=2,
            value=f"{hour.get('first_name', '')} {hour.get('last_name', '')}",
        )
        ws_detail.cell(row=row_idx, column=3, value=hour.get("duration_hours", 0))
        ws_detail.cell(row=row_idx, column=4, value=hour.get("source", "propre"))
        ws_detail.cell(row=row_idx, column=5, value=hour.get("reason", ""))
        ws_detail.cell(row=row_idx, column=6, value=hour.get("meeting_title", ""))
    for col_idx in range(1, len(headers_detail) + 1):
        h = headers_detail[col_idx - 1]
        max_length = max(
            len(str(h)),
            max(
                (
                    len(str(ws_detail.cell(row=row_idx, column=col_idx).value or ""))
                    for row_idx in range(2, ws_detail.max_row + 1)
                ),
                default=0,
            ),
        )
        ws_detail.column_dimensions[
            ws_detail.cell(row=1, column=col_idx).column_letter
        ].width = min(max_length + 2, 50)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def export_meetings_history(meetings: List[Dict[str, Any]]) -> bytes:
    """Export Excel de l'historique des réunions CSE."""
    headers = [
        "Titre",
        "Date",
        "Heure",
        "Type",
        "Statut",
        "Lieu",
        "Nombre participants",
        "PV généré",
    ]
    data = []
    for meeting in meetings:
        meeting_date = _formater_date(meeting.get("meeting_date"))
        meeting_time = _formater_heure(meeting.get("meeting_time"))
        status_labels = {
            "a_venir": "À venir",
            "en_cours": "En cours",
            "terminee": "Terminée",
        }
        status = status_labels.get(meeting.get("status", ""), meeting.get("status", ""))
        type_labels = {
            "ordinaire": "Ordinaire",
            "extraordinaire": "Extraordinaire",
            "cssct": "CSSCT",
            "autre": "Autre",
        }
        meeting_type = type_labels.get(
            meeting.get("meeting_type", ""), meeting.get("meeting_type", "")
        )
        has_minutes = "Oui" if meeting.get("has_minutes") else "Non"
        data.append(
            {
                "Titre": meeting.get("title", ""),
                "Date": meeting_date,
                "Heure": meeting_time,
                "Type": meeting_type,
                "Statut": status,
                "Lieu": meeting.get("location", ""),
                "Nombre participants": meeting.get("participant_count", 0),
                "PV généré": has_minutes,
            }
        )
    return generate_xlsx(data, headers, "Historique réunions CSE")
