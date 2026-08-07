"""Tests unitaires — détection colonnes Excel RIB."""

from io import BytesIO

import pytest

from app.modules.admin_import.application.rib_parser import parse_rib_cell

from app.modules.admin_import.application.rib_excel import (
    detect_rib_column_mapping,
    find_header_row_index,
    read_tabular_file,
    rib_cell_value,
)


class TestDetectRibColumnMapping:
    def test_detects_rib_and_identity_columns(self):
        headers = ["Nom", "Prénom", "Matricule", "RIB", "Email"]
        mapping = detect_rib_column_mapping(headers)
        assert mapping["rib"] == "RIB"
        assert mapping["last_name"] == "Nom"
        assert mapping["first_name"] == "Prénom"
        assert mapping["matricule"] == "Matricule"
        assert mapping["email"] == "Email"

    def test_detects_iban_alias(self):
        mapping = detect_rib_column_mapping(["Salarié", "IBAN"])
        assert mapping["rib"] == "IBAN"
        assert mapping["full_name"] == "Salarié"


class TestFindHeaderRowIndex:
    def test_finds_header_on_third_row(self):
        raw = [
            ["Export RIB salariés"],
            ["Entreprise Demo", "", ""],
            ["Nom", "Prénom", "RIB"],
            ["Martin", "Paul", "FR1420041010050500013M02606"],
        ]
        assert find_header_row_index(raw) == 2

    def test_ignores_title_containing_rib_word(self):
        raw = [
            ["Export RIB salariés janvier"],
            ["Nom", "Prénom", "RIB"],
            ["Martin", "Paul", "FR1420041010050500013M02606"],
        ]
        assert find_header_row_index(raw) == 1


class TestReadTabularFile:
    def test_reads_csv(self):
        content = (
            "Nom,Prénom,RIB\n"
            "Martin,Paul,FR1420041010050500013M02606\n"
        ).encode("utf-8")
        sheet = read_tabular_file(content, "ribs.csv")
        assert sheet.headers == ["Nom", "Prénom", "RIB"]
        assert len(sheet.rows) == 1
        assert sheet.rows[0]["Nom"] == "Martin"

    def test_reads_csv_with_preamble(self):
        content = (
            "Export RIB\n"
            "Société Test\n"
            "Nom;Prénom;RIB\n"
            "Martin;Paul;FR1420041010050500013M02606\n"
        ).encode("utf-8")
        sheet = read_tabular_file(content, "ribs.csv")
        assert sheet.headers == ["Nom", "Prénom", "RIB"]
        assert sheet.header_row_index == 3
        assert len(sheet.rows) == 1

    def test_reads_xlsx_with_header_on_third_row(self):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Export RIB salariés"
        ws["A2"] = "Entreprise XYZ"
        ws["A3"] = "Nom"
        ws["B3"] = "Prénom"
        ws["C3"] = "RIB"
        ws["A4"] = "Martin"
        ws["B4"] = "Paul"
        ws["C4"] = "FR1420041010050500013M02606"
        buf = BytesIO()
        wb.save(buf)

        sheet = read_tabular_file(buf.getvalue(), "ribs.xlsx")
        assert sheet.headers == ["Nom", "Prénom", "RIB"]
        assert sheet.header_row_index == 3
        assert len(sheet.rows) == 1
        assert sheet.rows[0]["Nom"] == "Martin"


class TestIbanEclateSurDeuxColonnes:
    """Export Quadratus « liste des employés » : l'IBAN tient sur deux colonnes.

    « Bq iban » ne porte que le pays et la clé (FR76), « Bq rib » les 23 caractères
    du BBAN. La colonne retenue seule ne vaut rien : c'est le recollage qui rend
    l'IBAN. Les sept fichiers d'Elsa du 27/07/2026 sont dans ce format, sauf LEWIS.
    """

    HEADERS = ["Numero", "Nom", "Bq iban", "Bq rib", "Bq bic"]

    def test_les_deux_colonnes_rib_sont_retenues(self):
        mapping = detect_rib_column_mapping(self.HEADERS)
        assert mapping["rib"] == "Bq iban"
        assert mapping["rib_complement"] == "Bq rib"
        assert mapping["bic"] == "Bq bic"

    def test_le_prefixe_et_le_bban_sont_recolles(self):
        mapping = detect_rib_column_mapping(self.HEADERS)
        row = {
            "Numero": "BUGNY",
            "Nom": "BUGNY",
            "Bq iban": "FR76",
            "Bq rib": "13825002000494171238574",
            "Bq bic": "CEPAFRPP382",
        }
        assert rib_cell_value(row, mapping) == "FR7613825002000494171238574"

    def test_le_recollage_donne_un_iban_valide(self):
        mapping = detect_rib_column_mapping(self.HEADERS)
        row = {"Bq iban": "FR76", "Bq rib": "13825002000494171238574", "Bq bic": "CEPAFRPP382"}
        iban, bic, valide, erreur = parse_rib_cell(
            rib_cell_value(row, mapping), bic_hint=row["Bq bic"]
        )
        assert valide, erreur
        assert iban == "FR7613825002000494171238574"
        assert bic == "CEPAFRPP382"

    def test_une_seule_colonne_rib_reste_inchangee(self):
        """Format LEWIS : une colonne « RIB » portant le BBAN seul. Rien à recoller."""
        mapping = detect_rib_column_mapping(["Nom", "RIB"])
        assert "rib_complement" not in mapping
        row = {"Nom": "BASTER", "RIB": "10278374050001248800154"}
        assert rib_cell_value(row, mapping) == "10278374050001248800154"

    def test_deux_colonnes_completes_ne_sont_pas_collees(self):
        """Garde-fou : deux coordonnées distinctes ne doivent jamais être concaténées.

        On ne recolle que si la première colonne est un préfixe seul (pays + clé).
        """
        mapping = detect_rib_column_mapping(["IBAN", "RIB"])
        row = {
            "IBAN": "FR7613825002000494171238574",
            "RIB": "17806002016225312669363",
        }
        assert rib_cell_value(row, mapping) == "FR7613825002000494171238574"

    def test_prefixe_seul_sans_complement_reste_tel_quel(self):
        mapping = detect_rib_column_mapping(["Bq iban", "Bq bic"])
        assert rib_cell_value({"Bq iban": "FR76", "Bq bic": ""}, mapping) == "FR76"
