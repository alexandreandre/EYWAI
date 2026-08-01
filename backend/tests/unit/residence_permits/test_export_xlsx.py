"""
Fabrication du fichier XLSX des titres de séjour.

Les tests relisent le classeur produit : ils valident ce qu'Elsa ouvrira dans Excel,
pas la structure intermédiaire. Le champ `residence_permit_type` vaut NULL pour les
43 salariés soumis en production ; une cellule vide, jamais « None », est donc le cas
nominal et non un cas limite.
"""

from __future__ import annotations

import io
from datetime import date

import pytest
from openpyxl import load_workbook

from app.modules.residence_permits.infrastructure.export_xlsx import (
    EXPORT_HEADERS,
    build_export_filename,
    build_residence_permits_xlsx,
)

pytestmark = pytest.mark.unit


def _row(**kwargs):
    base = {
        "id": "emp-1",
        "first_name": "Dieu Merci",
        "last_name": "LANKOKO MVUKI",
        "matricule": "000123",
        "job_title": "Opérateur",
        "hire_date": "2023-04-03",
        "nationalite": "CONGOLAISE",
        "employment_status": "actif",
        "residence_permit_status": "expired",
        "residence_permit_type": None,
        "residence_permit_number": "9912345678",
        "residence_permit_expiry_date": "2026-01-28",
        "residence_permit_days_remaining": -184,
    }
    base.update(kwargs)
    return base


def _sheet(rows, company_name="Mont Blanc Composite"):
    content = build_residence_permits_xlsx(rows, company_name)
    return load_workbook(io.BytesIO(content)).active


def _values(ws, row_idx):
    return [c.value for c in ws[row_idx]]


def test_entetes_dans_l_ordre():
    ws = _sheet([_row()])
    assert _values(ws, 1) == EXPORT_HEADERS
    assert len(EXPORT_HEADERS) == 13


def test_ligne_complete():
    """Le « Type de titre » se relit None : un tableur n'a pas de chaîne vide,
    seulement des cellules vides. C'est le cas nominal ici (NULL pour les 43)."""
    ws = _sheet([_row()])
    assert _values(ws, 2) == [
        "LANKOKO MVUKI",
        "Dieu Merci",
        "000123",
        "Mont Blanc Composite",
        "Opérateur",
        "03/04/2023",
        "CONGOLAISE",
        "Actif",
        "Expiré",
        None,
        "9912345678",
        "28/01/2026",
        -184,
    ]


def test_dates_au_format_francais():
    ws = _sheet(
        [_row(hire_date=date(2024, 12, 1), residence_permit_expiry_date="2027-02-09")]
    )
    ligne = _values(ws, 2)
    assert ligne[5] == "01/12/2024"
    assert ligne[11] == "09/02/2027"


@pytest.mark.parametrize(
    "statut,libelle",
    [
        ("expired", "Expiré"),
        ("to_renew", "À renouveler"),
        ("to_complete", "À compléter"),
        ("valid", "Valide"),
        (None, "À compléter"),
    ],
)
def test_libelles_de_statut(statut, libelle):
    ws = _sheet([_row(residence_permit_status=statut)])
    assert _values(ws, 2)[8] == libelle


def test_statut_emploi_en_toutes_lettres():
    ws = _sheet([_row(employment_status="en_sortie")])
    assert _values(ws, 2)[7] == "En sortie"


def test_valeurs_absentes_donnent_une_cellule_vide():
    """Ni « None », ni « — » : une cellule vide, qu'Excel sait filtrer et trier."""
    ws = _sheet(
        [
            _row(
                residence_permit_type=None,
                residence_permit_number=None,
                residence_permit_expiry_date=None,
                residence_permit_days_remaining=None,
                job_title=None,
                nationalite=None,
            )
        ]
    )
    ligne = _values(ws, 2)
    for index in (4, 6, 9, 10, 11, 12):
        assert ligne[index] in ("", None), f"colonne {index} = {ligne[index]!r}"


def test_jours_restants_negatif_pour_un_titre_expire():
    """Le nombre reste un nombre : Excel doit pouvoir trier par urgence."""
    ws = _sheet([_row(residence_permit_days_remaining=-184)])
    valeur = _values(ws, 2)[12]
    assert valeur == -184
    assert isinstance(valeur, int)


def test_ordre_des_lignes_preserve():
    ws = _sheet([_row(id="a", last_name="AAA"), _row(id="b", last_name="BBB")])
    assert _values(ws, 2)[0] == "AAA"
    assert _values(ws, 3)[0] == "BBB"


def test_aucune_ligne_produit_un_classeur_avec_entetes():
    ws = _sheet([])
    assert _values(ws, 1) == EXPORT_HEADERS
    assert ws.max_row == 1


def test_nom_de_fichier():
    assert (
        build_export_filename("Mont Blanc Composite", date(2026, 7, 31))
        == "titres-de-sejour_mont-blanc-composite_2026-07-31.xlsx"
    )


def test_nom_de_fichier_accents_et_ponctuation():
    assert (
        build_export_filename("Cartol Industrie (S.A.)", date(2026, 7, 31))
        == "titres-de-sejour_cartol-industrie-s-a_2026-07-31.xlsx"
    )


def test_nom_de_fichier_societe_vide():
    assert (
        build_export_filename("", date(2026, 7, 31))
        == "titres-de-sejour_entreprise_2026-07-31.xlsx"
    )
