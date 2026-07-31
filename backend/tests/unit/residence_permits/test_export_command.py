"""
Cas d'usage d'export : garde-fous et restauration de l'ordre demandé.

Le navigateur envoie les identifiants dans l'ordre d'affichage (tri par urgence).
PostgREST ne garantit aucun ordre sur un `IN` : sans restauration explicite, le
fichier ne correspondrait plus à l'écran, ce qui est toute la promesse de l'export.
"""

from __future__ import annotations

import io
from datetime import date

import pytest
from openpyxl import load_workbook

from app.modules.residence_permits.application.exports import (
    MAX_EXPORT_EMPLOYEES,
    ResidencePermitExportEmpty,
    ResidencePermitExportTooLarge,
    export_residence_permits,
)

pytestmark = pytest.mark.unit

COMPANY_ID = "550e8400-e29b-41d4-a716-446655440000"


class FakeReader:
    """Simule la borne serveur : ne rend que ce qui appartient à l'entreprise."""

    def __init__(self, rows_par_entreprise):
        self.rows_par_entreprise = rows_par_entreprise
        self.appels = []

    def get_employees_for_export(self, company_id, employee_ids):
        self.appels.append((company_id, list(employee_ids)))
        connus = self.rows_par_entreprise.get(company_id, {})
        # Ordre volontairement inversé : PostgREST ne garantit rien.
        return [connus[i] for i in reversed(employee_ids) if i in connus]


def _row(emp_id, last_name, **kwargs):
    base = {
        "id": emp_id,
        "first_name": "Test",
        "last_name": last_name,
        "matricule": "0001",
        "job_title": "Opérateur",
        "hire_date": "2023-01-02",
        "nationalite": "FRANCAISE",
        "employment_status": "actif",
        "is_subject_to_residence_permit": True,
        "residence_permit_expiry_date": "2027-01-01",
        "residence_permit_type": None,
        "residence_permit_number": "123",
    }
    base.update(kwargs)
    return base


def _noms(content):
    ws = load_workbook(io.BytesIO(content)).active
    return [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]


def test_ordre_demande_restaure():
    """Le fichier suit l'ordre du navigateur, pas celui de la base."""
    reader = FakeReader(
        {
            COMPANY_ID: {
                "a": _row("a", "AAA"),
                "b": _row("b", "BBB"),
                "c": _row("c", "CCC"),
            }
        }
    )
    content, _ = export_residence_permits(
        COMPANY_ID, "Test Co", ["a", "b", "c"], reader=reader
    )

    assert _noms(content) == ["AAA", "BBB", "CCC"]


def test_identifiant_inconnu_ignore_sans_erreur():
    """Un salarié sorti entre l'affichage et le clic ne doit pas casser l'export."""
    reader = FakeReader({COMPANY_ID: {"a": _row("a", "AAA")}})
    content, _ = export_residence_permits(
        COMPANY_ID, "Test Co", ["a", "inexistant"], reader=reader
    )

    assert _noms(content) == ["AAA"]


def test_identifiant_d_une_autre_societe_exclu():
    """Cloisonnement : le lecteur n'est interrogé que sur l'entreprise active."""
    autre = "660e8400-e29b-41d4-a716-446655440099"
    reader = FakeReader(
        {COMPANY_ID: {"a": _row("a", "AAA")}, autre: {"z": _row("z", "ZZZ")}}
    )
    content, _ = export_residence_permits(
        COMPANY_ID, "Test Co", ["a", "z"], reader=reader
    )

    assert _noms(content) == ["AAA"]
    assert reader.appels == [(COMPANY_ID, ["a", "z"])]


def test_liste_vide_refusee():
    reader = FakeReader({COMPANY_ID: {}})
    with pytest.raises(ResidencePermitExportEmpty):
        export_residence_permits(COMPANY_ID, "Test Co", [], reader=reader)


def test_aucune_correspondance_refusee():
    reader = FakeReader({COMPANY_ID: {}})
    with pytest.raises(ResidencePermitExportEmpty):
        export_residence_permits(COMPANY_ID, "Test Co", ["inexistant"], reader=reader)


def test_trop_d_identifiants_refuse():
    reader = FakeReader({COMPANY_ID: {}})
    trop = [f"emp-{i}" for i in range(MAX_EXPORT_EMPLOYEES + 1)]
    with pytest.raises(ResidencePermitExportTooLarge):
        export_residence_permits(COMPANY_ID, "Test Co", trop, reader=reader)


def test_doublons_dedupliques():
    reader = FakeReader({COMPANY_ID: {"a": _row("a", "AAA")}})
    content, _ = export_residence_permits(
        COMPANY_ID, "Test Co", ["a", "a"], reader=reader
    )

    assert _noms(content) == ["AAA"]
    assert reader.appels == [(COMPANY_ID, ["a"])]


def test_statut_calcule_present_dans_le_fichier():
    """L'enrichissement est fait ici : le fichier porte un statut, pas un champ brut."""
    reader = FakeReader(
        {COMPANY_ID: {"a": _row("a", "AAA", residence_permit_expiry_date="2020-01-01")}}
    )
    content, _ = export_residence_permits(COMPANY_ID, "Test Co", ["a"], reader=reader)
    ws = load_workbook(io.BytesIO(content)).active

    assert ws.cell(row=2, column=9).value == "Expiré"


def test_nom_de_fichier_retourne():
    reader = FakeReader({COMPANY_ID: {"a": _row("a", "AAA")}})
    _, filename = export_residence_permits(
        COMPANY_ID,
        "Mont Blanc Composite",
        ["a"],
        reader=reader,
        today=date(2026, 7, 31),
    )

    assert filename == "titres-de-sejour_mont-blanc-composite_2026-07-31.xlsx"
