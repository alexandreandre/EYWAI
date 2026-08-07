"""L'export des élus reçoit des objets date (model_dump), pas des chaînes ISO.

Avant correction, datetime.fromisoformat(objet_date) levait TypeError, l'exception
était avalée, et tout mandat expiré sortait « Actif » avec « Jours restants » vide.
"""

import io
from datetime import date, timedelta

import openpyxl

from app.modules.cse.infrastructure.cse_export_impl import export_elected_members


def _membre(fin: date, days_remaining=None) -> dict:
    return {
        "id": "elu-1",
        "employee_id": "emp-1",
        "first_name": "Prenom",
        "last_name": "NOM",
        "job_title": "Opérateur",
        "role": "titulaire",
        "college": "1er collège",
        "start_date": fin - timedelta(days=1461),
        "end_date": fin,
        "is_active": True,
        "days_remaining": days_remaining,
    }


def _lignes(contenu: bytes) -> list[dict]:
    ws = openpyxl.load_workbook(io.BytesIO(contenu)).active
    entetes = [c.value for c in ws[1]]
    return [dict(zip(entetes, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]


def test_mandat_expire_sort_en_statut_expire():
    contenu = export_elected_members([_membre(date.today() - timedelta(days=45))])
    ligne = _lignes(contenu)[0]
    assert ligne["Statut"] == "Expiré"
    assert ligne["Jours restants"] == -45


def test_mandat_qui_expire_bientot():
    contenu = export_elected_members([_membre(date.today() + timedelta(days=30))])
    assert _lignes(contenu)[0]["Statut"] == "Expire bientôt"


def test_mandat_en_cours_reste_actif():
    contenu = export_elected_members([_membre(date.today() + timedelta(days=400))])
    assert _lignes(contenu)[0]["Statut"] == "Actif"


def test_les_dates_sortent_au_format_francais():
    contenu = export_elected_members([_membre(date(2027, 3, 9))])
    ligne = _lignes(contenu)[0]
    assert ligne["Date fin mandat"] == "09/03/2027"
    assert ligne["Date début mandat"] == "09/03/2023"


def test_accepte_aussi_des_chaines_iso():
    membre = _membre(date.today() - timedelta(days=45))
    membre["end_date"] = membre["end_date"].isoformat()
    membre["start_date"] = membre["start_date"].isoformat()
    ligne = _lignes(export_elected_members([membre]))[0]
    assert ligne["Statut"] == "Expiré"


def test_sans_date_de_fin_le_statut_est_inconnu():
    membre = _membre(date.today())
    membre["end_date"] = None
    ligne = _lignes(export_elected_members([membre]))[0]
    assert ligne["Statut"] == "Inconnu"
    # openpyxl convertit une cellule "" en None au round-trip save/load : la valeur
    # écrite est bien "" (cf. export_elected_members), mais relue elle vaut None.
    assert ligne["Jours restants"] is None


def test_mandat_revoque_sort_revoque_meme_si_la_date_de_fin_nest_pas_atteinte():
    membre = _membre(date.today() + timedelta(days=400))
    membre["is_active"] = False
    ligne = _lignes(export_elected_members([membre]))[0]
    assert ligne["Statut"] == "Révoqué"


def test_absence_de_is_active_ne_regresse_pas_vers_revoque():
    membre = _membre(date.today() + timedelta(days=400))
    del membre["is_active"]
    ligne = _lignes(export_elected_members([membre]))[0]
    assert ligne["Statut"] == "Actif"


def test_mandat_expire_avec_is_active_faux_reste_expire():
    """Cas du bug : l'import pose is_active=False *parce que* la date de fin est
    passée. La date de fin doit primer : c'est « Expiré », pas « Révoqué »."""
    membre = _membre(date.today() - timedelta(days=45))
    membre["is_active"] = False
    ligne = _lignes(export_elected_members([membre]))[0]
    assert ligne["Statut"] == "Expiré"
