"""Les exports heures et réunions reçoivent eux aussi des objets date (model_dump)."""

import io
from datetime import date, time

import openpyxl

from app.modules.cse.infrastructure.cse_export_impl import (
    export_delegation_hours,
    export_meetings_history,
)


def _cellule(contenu: bytes, feuille: str, ligne: int, colonne: int):
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    return wb[feuille].cell(row=ligne, column=colonne).value


def test_detail_des_heures_date_au_format_francais():
    heures = [
        {
            "date": date(2026, 7, 9),
            "first_name": "Prenom",
            "last_name": "NOM",
            "duration_hours": 4,
            "source": "propre",
            "reason": "Réunion",
            "meeting_title": "CSE juillet",
        }
    ]
    contenu = export_delegation_hours(heures, [])
    assert _cellule(contenu, "Détail heures", 2, 1) == "09/07/2026"


def test_historique_des_reunions_date_et_heure_au_format_francais():
    # meeting_time est un vrai datetime.time en production (model_dump en mode python) :
    # une chaîne ici certifierait la colonne Date mais déguiserait un bug sur l'Heure
    # (time[:5] lève TypeError, avalée par un except Exception: pass avant correction).
    reunions = [
        {
            "title": "CSE juillet",
            "meeting_date": date(2026, 7, 9),
            "meeting_time": time(14, 0, 0),
            "meeting_type": "ordinaire",
            "status": "terminee",
            "location": "Salle A",
            "participant_count": 4,
            "has_minutes": True,
        }
    ]
    contenu = export_meetings_history(reunions)
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb.active
    entetes = [c.value for c in ws[1]]
    ligne = dict(zip(entetes, [c.value for c in ws[2]]))
    assert ligne["Date"] == "09/07/2026"
    assert ligne["Heure"] == "14:00"
