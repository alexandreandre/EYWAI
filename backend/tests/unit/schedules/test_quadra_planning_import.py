"""Tests parseur calendrier prévu Excel Quadra."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
import pytest

from app.modules.schedules.application.employee_match import (
    resolve_employee_for_planning_sheet,
)
from app.modules.schedules.application.planning_import.quadra_calendar import (
    classify_planned_day,
    is_quadra_planning_workbook,
    parse_quadra_planning_workbook,
)
from app.modules.schedules.application.timesheet_import.tabular_period import (
    ImportPeriodConfig,
)
from app.modules.schedules.schemas.ai import RosterEmployee

XLSX = Path(
    "/Users/alex/Desktop/Comitech Composite/Calendrier/calendrier 2026 comitech.xlsx"
)

ROSTER = [
    RosterEmployee(id="1", first_name="Samir", last_name="BOUFRIDA"),
    RosterEmployee(id="2", first_name="Michel", last_name="BOUVEYRON"),
    RosterEmployee(
        id="3",
        first_name="Vitor Manuel",
        last_name="CASANOVA DA SILVA",
    ),
    RosterEmployee(id="4", first_name="Lucas", last_name="CHAMBERT"),
]


@pytest.mark.skipif(not XLSX.is_file(), reason="Fichier calendrier Comitech absent")
def test_detect_quadra_workbook() -> None:
    content = XLSX.read_bytes()
    assert is_quadra_planning_workbook(content, XLSX.name)


@pytest.mark.skipif(not XLSX.is_file(), reason="Fichier calendrier Comitech absent")
def test_parse_quadra_year_mode() -> None:
    content = XLSX.read_bytes()
    parsed = parse_quadra_planning_workbook(
        content,
        XLSX.name,
        year=2026,
        period_config=ImportPeriodConfig(mode="year", year=2026, month=1),
        roster=ROSTER,
    )
    assert parsed.sheets_parsed >= 10
    assert len(parsed.month_groups) == 12
    first_group = parsed.month_groups[0]
    assert first_group["year"] == 2026
    assert first_group["month"] == 1
    employees = first_group["employees"]
    assert employees
    assert all(day["nature"] == "prevu" for emp in employees for day in emp["days"])


def test_resolve_sheet_last_name_only() -> None:
    match = resolve_employee_for_planning_sheet("BOUFRIDA", ROSTER)
    assert match.employee_id == "1"
    assert match.review_status == "ok"


def test_resolve_sheet_with_sommaire_hint_disambiguates() -> None:
    roster = [
        RosterEmployee(id="3", first_name="Vitor Manuel", last_name="CASANOVA DA SILVA"),
        RosterEmployee(id="5", first_name="Vitor Manuel", last_name="DA SILVA CARDOSO"),
        RosterEmployee(id="2", first_name="Michel", last_name="BOUVEYRON"),
    ]
    match = resolve_employee_for_planning_sheet(
        "CASANOVA",
        roster,
        hint_name="CASANOVA DA SILVA Vitor Manuel",
    )
    assert match.employee_id == "3"
    assert match.review_status in ("ok", "warning")

    match_b = resolve_employee_for_planning_sheet(
        "BOUVEYRON",
        roster,
        hint_name="BOUVEYRON Michel",
    )
    assert match_b.employee_id == "2"
    assert match_b.review_status in ("ok", "warning")


def test_resolve_cartol_sheet_with_first_name_initials() -> None:
    roster = [
        RosterEmployee(id="jm", first_name="Jean-Michel", last_name="BONNET"),
        RosterEmployee(id="m", first_name="Mathieu", last_name="BONNET"),
        RosterEmployee(id="jn", first_name="Jean-Noël", last_name="LEMAIRE"),
        RosterEmployee(id="jo", first_name="Jordan", last_name="LEMAIRE"),
        RosterEmployee(id="gm", first_name="Guy-Marie", last_name="RAINGEAUD"),
        RosterEmployee(id="gio", first_name="Giovanni", last_name="RAINGEAUD"),
    ]

    for raw, expected in (
        ("BONNET JM", "jm"),
        ("LEMAIRE JN", "jn"),
        ("LEMAIRE JO", "jo"),
        ("RAINGEAUD G-M", "gm"),
        ("RAINGEAUD Gio", "gio"),
    ):
        match = resolve_employee_for_planning_sheet(raw, roster)
        assert match.employee_id == expected
        assert match.review_status == "ok"


def test_resolve_cartol_sheet_with_sommaire_initial_hint() -> None:
    roster = [
        RosterEmployee(id="s", first_name="Sulivan", last_name="CAILLEAU"),
        RosterEmployee(id="g", first_name="Gwendoline", last_name="CAILLEAUX"),
        RosterEmployee(id="o", first_name="Oleksandr", last_name="DOVHOPOL"),
        RosterEmployee(id="t", first_name="Tetiana", last_name="DOVHOPOL"),
    ]

    assert (
        resolve_employee_for_planning_sheet(
            "CAILLEAU",
            roster,
            hint_name="CAILLEAU S",
        ).employee_id
        == "s"
    )
    assert (
        resolve_employee_for_planning_sheet(
            "CAILLEAUX",
            roster,
            hint_name="CAILLEAUX G",
        ).employee_id
        == "g"
    )
    assert (
        resolve_employee_for_planning_sheet(
            "DOVHOPOL",
            roster,
            hint_name="DOVHOPOL T",
        ).employee_id
        == "t"
    )


def test_resolve_cartol_particle_last_names_are_not_first_name_hints() -> None:
    roster = [
        RosterEmployee(id="de-sa", first_name="Anthony", last_name="DE SA"),
        RosterEmployee(id="de-abreu", first_name="Jose", last_name="DE ABREU"),
        RosterEmployee(id="de-carvalho", first_name="Roberto", last_name="DE CARVALHO"),
        RosterEmployee(id="dovhopol", first_name="Oleksandr", last_name="DOVHOPOL"),
    ]

    match = resolve_employee_for_planning_sheet(
        "DE SA",
        roster,
        hint_name="DE SA",
    )
    assert match.employee_id == "de-sa"
    assert match.review_status == "ok"

    match = resolve_employee_for_planning_sheet(
        "DE ABREU",
        roster,
        hint_name="DE ABREU",
    )
    assert match.employee_id == "de-abreu"
    assert match.review_status == "ok"

    match = resolve_employee_for_planning_sheet(
        "DE CARVALHO",
        roster,
        hint_name="De CARVALHO",
    )
    assert match.employee_id == "de-carvalho"
    assert match.review_status == "ok"


def test_classify_cp_day() -> None:
    from datetime import date

    entry = classify_planned_day(date(2026, 2, 3), None, 1.0)
    assert entry["type"] == "conge"
    assert entry["heures_prevues"] == pytest.approx(7.8)


def test_parse_sheet_with_employee_metadata_row() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BASTER"
    ws.cell(1, 1).value = "Date d'ancienneté:"
    ws.cell(1, 8).value = "BASTER Damien"
    ws.cell(2, 1).value = "JANVIER"
    ws.cell(2, 3).value = "H.Abs"
    ws.cell(2, 4).value = "CP"
    ws.cell(3, 1).value = "jeudi"
    ws.cell(3, 2).value = 1
    ws.cell(3, 3).value = "jour de l'an "
    ws.cell(4, 1).value = "vendredi"
    ws.cell(4, 2).value = 2

    content = _workbook_bytes(wb)
    parsed = parse_quadra_planning_workbook(
        content,
        "calendrier 2026 LEWIS.xlsx",
        year=2026,
        period_config=ImportPeriodConfig(mode="month", year=2026, month=1),
        roster=[RosterEmployee(id="1", first_name="Damien", last_name="BASTER")],
    )

    assert parsed.sheets_parsed == 1
    employees = parsed.month_groups[0]["employees"]
    assert employees[0]["employee_id"] == "1"
    assert employees[0]["days"][0]["jour"] == 1
    assert employees[0]["days"][0]["type"] == "ferie"


def test_parse_cartol_sheet_uses_employee_header_to_disambiguate() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "COUTANT"
    ws.cell(1, 1).value = "Date d'ancienneté: 15/12/2014"
    ws.cell(1, 10).value = "COUTANT Denis"
    ws.cell(2, 1).value = "JANVIER"
    ws.cell(2, 3).value = "H.Abs"
    ws.cell(2, 4).value = "CP"
    ws.cell(3, 1).value = "jeudi"
    ws.cell(3, 2).value = 1
    ws.cell(3, 3).value = "jour de l'an "

    parsed = parse_quadra_planning_workbook(
        _workbook_bytes(wb),
        "calendrier 2026 CARTOL.xlsx",
        year=2026,
        period_config=ImportPeriodConfig(mode="month", year=2026, month=1),
        roster=[
            RosterEmployee(id="denis", first_name="Denis", last_name="COUTANT"),
            RosterEmployee(id="guillaume", first_name="Guillaume", last_name="COUTANT"),
        ],
    )

    employee = parsed.month_groups[0]["employees"][0]
    assert employee["employee_id"] == "denis"
    assert employee["review_status"] == "ok"
    assert employee["sommaire_hint"] == "COUTANT Denis"


def test_resolve_planning_hint_does_not_match_first_name_only() -> None:
    match = resolve_employee_for_planning_sheet(
        "PEROT",
        [RosterEmployee(id="veillat", first_name="Sébastien", last_name="VEILLAT")],
        hint_name="PEROT Sébastien",
    )

    assert match.employee_id is None
    assert match.review_status == "error"


def test_resolve_planning_old_compound_last_name_by_unique_first_name() -> None:
    match = resolve_employee_for_planning_sheet(
        "ESPIRITO SANTO",
        [RosterEmployee(id="dias", first_name="Tania", last_name="DIAS")],
        hint_name="ESPIRITO SANTO Tania",
    )

    assert match.employee_id == "dias"
    assert match.review_status == "ok"


def test_resolve_planning_partial_compound_last_name_with_first_hint_is_ok() -> None:
    match = resolve_employee_for_planning_sheet(
        "ZAROUALI",
        [
            RosterEmployee(
                id="zarouali",
                first_name="EL HOUSINE",
                last_name="ZAROUALI BOUTABAA",
            )
        ],
        hint_name="ZAROUALI El Houcine",
    )

    assert match.employee_id == "zarouali"
    assert match.review_status == "ok"


def test_resolve_planning_simple_last_name_does_not_match_first_name_only() -> None:
    match = resolve_employee_for_planning_sheet(
        "DEPLANNE",
        [RosterEmployee(id="enond", first_name="Marie-Noëlle", last_name="ENOND")],
        hint_name="DEPLANNE Marie-Noëlle",
    )

    assert match.employee_id is None
    assert match.review_status == "error"


def test_resolve_planning_sheet_matches_hyphenated_last_name() -> None:
    match = resolve_employee_for_planning_sheet(
        "SELLY-PAJADON",
        [RosterEmployee(id="selly", first_name="Sandy", last_name="SELLY PAJADON")],
        hint_name="SELLY-PAJADON Sandy",
    )

    assert match.employee_id == "selly"
    assert match.review_status == "ok"


def _workbook_bytes(wb: Workbook) -> bytes:
    from io import BytesIO

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
