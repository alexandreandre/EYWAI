"""Parse le calendrier Excel Quadra Comitech Composite → calendrier_prevu EYWAI."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from app.modules.schedules.application.planning_import.quadra_calendar import (
    build_default_month_calendar,
    classify_planned_day,
    parse_employee_sheet,
)
from scripts.comitech_calendar_data import (
    CALENDAR_SOURCE,
    COMITECH_CALENDAR_SHEETS,
    CalendarSheetMapping,
    DEFAULT_XLSX,
)

SKIP_SHEETS = frozenset({"SOMMAIRE", "MODELE"})


def _sheet_key(name: str) -> str:
    from app.modules.schedules.application.planning_import.quadra_calendar import (
        _sheet_key as key,
    )

    return key(name)


def load_workbook_sheets(xlsx_path: Path | None = None) -> dict[str, Any]:
    path = xlsx_path or DEFAULT_XLSX
    if not path.is_file():
        raise FileNotFoundError(f"Calendrier Excel introuvable : {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets: dict[str, Any] = {}
    for name in wb.sheetnames:
        key = _sheet_key(name)
        if key in SKIP_SHEETS:
            continue
        sheets[key] = wb[name]
    return sheets


def sheet_mapping_by_key() -> dict[str, CalendarSheetMapping]:
    return {_sheet_key(m.sheet_key): m for m in COMITECH_CALENDAR_SHEETS}


def build_planned_calendar_payload(
    entries: list[dict[str, Any]],
    *,
    year: int,
    month: int,
) -> dict[str, Any]:
    return {
        "periode": {"mois": month, "annee": year},
        "source": CALENDAR_SOURCE,
        "calendrier_prevu": entries,
    }


__all__ = [
    "build_default_month_calendar",
    "build_planned_calendar_payload",
    "classify_planned_day",
    "load_workbook_sheets",
    "parse_employee_sheet",
    "sheet_mapping_by_key",
]
