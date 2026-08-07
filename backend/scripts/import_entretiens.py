"""Reprise des entretiens annuels depuis le classeur transmis par Elsa le 27/07/2026.

Le classeur porte deux choses distinctes :

1. La **politique d'entretien de chaque société** (colonne « Règle appliquée ») — un
   mois de campagne commun, ou la date d'ancienneté, et parfois un cycle de deux ans.
   Elle est écrite dans `company_interview_settings`, pas dans ce script : la RH la
   modifie ensuite sans nous, et EYWAI propose seul la campagne des années suivantes.

2. Le **dernier entretien connu** (colonne « Dernier entretien »), qui ne contient
   qu'une année et seulement pour une société. Il est repris tel quel : `year` renseigné,
   `completed_date` laissé vide. On ne fabrique pas un jour qu'on n'a pas.

L'échéance de chaque salarié est recalculée par le domaine (`domain/campaign.py`) et
comparée à la colonne « Entretien à planifier » du classeur : toute divergence bloque
l'écriture, sauf --ignorer-divergences. C'est ce qui garantit que le moteur et le
classeur disent la même chose avant qu'on écrive quoi que ce soit.

Aucun nom n'est écrit dans ce fichier : le dépôt est public, les données nominatives
restent sous data/.

Usage :
    python scripts/import_entretiens.py --fichier <classeur.xlsx> --dry-run
    python scripts/import_entretiens.py --fichier <classeur.xlsx> --apply \\
        --confirmer-production
"""

from __future__ import annotations

import argparse
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.modules.annual_reviews.domain.campaign import (  # noqa: E402
    InterviewCampaignSettings,
    deduce_interview_type,
    next_campaign_date,
)

# Le classeur nomme les sociétés autrement que la base.
SOCIETES = {
    "CARTOL": "Cartol Industrie",
    "LEWIS": "LEWIS",
    "MBC": "Mont Blanc Composite",
    "MONT BLANC COMPOSITE": "Mont Blanc Composite",
    "COLORPLAST": "Colorplast",
    "COMITECH": "Comitech Composite",
    "MAJI": "MAJI",
    "ZONE": "Zone 404 Mars",
    "ZONE 404": "Zone 404 Mars",
}

# Les huit libellés de la colonne « Règle appliquée », traduits en réglage société.
# None = la ligne ne porte aucune politique (elle constate seulement une absence
# d'historique) : elle ne participe pas à la déduction du réglage.
REGLES: Dict[str, Optional[Tuple[str, Optional[int], int]]] = {
    "Tous les entretiens à refaire en novembre 2026": ("mois_fixe", 11, 1),
    "Reprise : tous les entretiens à refaire en octobre": ("mois_fixe", 10, 1),
    "Entretien annuel chaque année en octobre": ("mois_fixe", 10, 1),
    "Tous les entretiens en décembre": ("mois_fixe", 12, 1),
    "Dernier entretien 2024 -> +2 ans (octobre)": ("mois_fixe", 10, 2),
    "Dernier entretien 2025 -> +2 ans (octobre)": ("mois_fixe", 10, 2),
    "À la date d'ancienneté de 1 an (entrée + 1 an)": ("anniversaire_embauche", None, 1),
    "Aucun entretien enregistré -> à planifier au plus tôt": None,
}

MOIS_FR = {
    "janvier": 1, "février": 2, "fevrier": 2, "mars": 3, "avril": 4, "mai": 5,
    "juin": 6, "juillet": 7, "août": 8, "aout": 8, "septembre": 9, "octobre": 10,
    "novembre": 11, "décembre": 12, "decembre": 12,
}

STATUTS_ELIGIBLES = frozenset({"actif", "active", "en_sortie", "en_onboarding"})

PROJECT_REF_PRODUCTION = "slleauhyjnmiawosvlcg"
DRAPEAU_CONFIRMATION_PRODUCTION = "--confirmer-production"
LIMITE_POSTGREST = 1000

SOURCE = "planif_entretiens_2026-07-27"
NOTE_REPRISE = (
    "Reprise du classeur transmis le 27/07/2026 : entretien tenu en {annee}, "
    "date exacte non communiquée."
)


@dataclass
class LigneClasseur:
    societe: str
    nom: str
    prenom: str
    date_entree: Optional[date]
    dernier_entretien: Optional[int]
    a_planifier_texte: str
    regle: str


@dataclass
class Bilan:
    reglages: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    historiques: List[Dict[str, Any]] = field(default_factory=list)
    planifications: List[Dict[str, Any]] = field(default_factory=list)
    ignores: List[Tuple[str, str]] = field(default_factory=list)
    bloques: List[Tuple[str, str]] = field(default_factory=list)
    divergences: List[Tuple[str, str]] = field(default_factory=list)


def cle_nom(valeur: Optional[str]) -> str:
    """Majuscules, sans accent, sans tiret ni espace — pour comparer deux noms."""
    sans_accent = unicodedata.normalize("NFD", (valeur or "").upper())
    sans_accent = sans_accent.encode("ascii", "ignore").decode()
    return sans_accent.replace("-", "").replace("'", "").replace(" ", "").strip()


def _vers_date(valeur: Any) -> Optional[date]:
    if valeur in (None, ""):
        return None
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    texte = str(valeur).strip()
    for motif in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(texte, motif).date()
        except ValueError:
            continue
    return None


def _vers_annee(valeur: Any) -> Optional[int]:
    """« 2024 » donne 2024 ; « Aucun » et le vide donnent None."""
    texte = str(valeur or "").strip()
    return int(texte) if texte.isdigit() and len(texte) == 4 else None


def lire_classeur(chemin: Path) -> List[LigneClasseur]:
    import openpyxl

    ws = openpyxl.load_workbook(chemin, data_only=True).active
    entetes = [str(c.value or "").strip() for c in ws[1]]
    lignes: List[LigneClasseur] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        valeurs = dict(zip(entetes, row))
        if not valeurs.get("Nom"):
            continue
        lignes.append(
            LigneClasseur(
                societe=str(valeurs.get("Société") or "").strip(),
                nom=str(valeurs.get("Nom") or "").strip(),
                prenom=str(valeurs.get("Prénom") or "").strip(),
                date_entree=_vers_date(valeurs.get("Date d'entrée")),
                dernier_entretien=_vers_annee(valeurs.get("Dernier entretien")),
                a_planifier_texte=str(valeurs.get("Entretien à planifier") or "").strip(),
                regle=str(valeurs.get("Règle appliquée") or "").strip(),
            )
        )
    return lignes


def deduire_reglages(
    lignes: List[LigneClasseur],
) -> Tuple[Dict[str, InterviewCampaignSettings], List[Tuple[str, str]]]:
    """Une politique par société, déduite de la colonne « Règle appliquée ».

    Deux règles incompatibles dans la même société arrêtent le script : mieux vaut
    demander à Elsa que d'en choisir une au hasard.
    """
    par_societe: Dict[str, Set[Tuple[str, Optional[int], int]]] = defaultdict(set)
    bloques: List[Tuple[str, str]] = []

    for ligne in lignes:
        if ligne.regle not in REGLES:
            bloques.append(
                (ligne.societe, f"règle inconnue dans le classeur : « {ligne.regle} »")
            )
            continue
        politique = REGLES[ligne.regle]
        if politique is not None:
            par_societe[ligne.societe].add(politique)

    reglages: Dict[str, InterviewCampaignSettings] = {}
    for societe, politiques in par_societe.items():
        if len(politiques) > 1:
            bloques.append(
                (societe, f"{len(politiques)} politiques différentes dans le classeur")
            )
            continue
        mode, mois, periodicite = next(iter(politiques))
        reglages[societe] = InterviewCampaignSettings(
            enabled=True,
            campaign_mode=mode,
            campaign_month=mois,
            periodicity_years=periodicite,
        )
    return reglages, bloques


def echeance_du_classeur(texte: str) -> Optional[date]:
    """« Octobre 2026 » ou « 05/01/2027 » — la précision varie selon la société."""
    exacte = _vers_date(texte)
    if exacte is not None:
        return exacte
    morceaux = texte.split()
    if len(morceaux) == 2 and morceaux[0].lower() in MOIS_FR:
        try:
            return date(int(morceaux[1]), MOIS_FR[morceaux[0].lower()], 1)
        except ValueError:
            return None
    return None


def _memes_echeances(calculee: date, classeur: Optional[date]) -> bool:
    """Le classeur donne un mois pour six sociétés, un jour précis pour la septième."""
    if classeur is None:
        return False
    if classeur.day == 1:
        return (calculee.year, calculee.month) == (classeur.year, classeur.month)
    return calculee == classeur


def planifier(
    lignes: List[LigneClasseur],
    reglages: Dict[str, InterviewCampaignSettings],
    societes: Dict[str, str],
    salaries: List[Dict[str, Any]],
    existants: Set[Tuple[str, int, str]],
    aujourdhui: date,
) -> Bilan:
    bilan = Bilan()

    index: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for salarie in salaries:
        for nom in (salarie.get("last_name"), salarie.get("nom_usage")):
            if nom:
                cle = (
                    salarie["company_id"],
                    cle_nom(nom),
                    cle_nom(salarie.get("first_name")),
                )
                index[cle].append(salarie)

    for societe, reglage in reglages.items():
        company_id = societes.get(SOCIETES.get(societe.upper(), ""))
        if company_id is None:
            bilan.bloques.append((societe, "société inconnue en base"))
            continue
        bilan.reglages[company_id] = {
            "enabled": reglage.enabled,
            "campaign_mode": reglage.campaign_mode,
            "campaign_month": reglage.campaign_month,
            "periodicity_years": reglage.periodicity_years,
        }

    for ligne in lignes:
        etiquette = f"{ligne.societe} {ligne.nom} {ligne.prenom}"
        nom_base = SOCIETES.get(ligne.societe.upper())
        company_id = societes.get(nom_base) if nom_base else None
        if company_id is None:
            bilan.bloques.append((etiquette, "société inconnue en base"))
            continue

        trouves = index.get(
            (company_id, cle_nom(ligne.nom), cle_nom(ligne.prenom)), []
        )
        if not trouves:
            bilan.ignores.append((etiquette, "introuvable parmi les salariés"))
            continue
        if len(trouves) > 1:
            bilan.bloques.append((etiquette, f"{len(trouves)} salariés homonymes"))
            continue

        salarie = trouves[0]
        if (salarie.get("employment_status") or "") not in STATUTS_ELIGIBLES:
            bilan.ignores.append(
                (etiquette, f"salarié {salarie.get('employment_status')}")
            )
            continue

        reglage = reglages.get(ligne.societe)
        if reglage is None:
            bilan.bloques.append((etiquette, "aucune politique déduite pour la société"))
            continue

        employee_id = salarie["id"]
        interview_type = deduce_interview_type(
            salarie.get("statut"), salarie.get("is_forfait_jour")
        )

        if ligne.dernier_entretien is not None:
            cle_historique = (employee_id, ligne.dernier_entretien, interview_type)
            if cle_historique not in existants:
                bilan.historiques.append(
                    {
                        "employee_id": employee_id,
                        "company_id": company_id,
                        "year": ligne.dernier_entretien,
                        "status": "realise",
                        "planned_date": None,
                        "completed_date": None,
                        "interview_type": interview_type,
                        "import_source": SOURCE,
                        "rh_notes": NOTE_REPRISE.format(annee=ligne.dernier_entretien),
                    }
                )

        # L'échéance vient du moteur, pas du classeur : c'est elle qui fera foi les
        # années suivantes, autant vérifier tout de suite qu'elles concordent.
        echeance = next_campaign_date(
            reglage,
            hire_date=_vers_date(salarie.get("hire_date")) or ligne.date_entree,
            last_review_year=ligne.dernier_entretien,
            today=aujourdhui,
        )
        if echeance is None:
            bilan.bloques.append((etiquette, "échéance non calculable (date d'entrée ?)"))
            continue

        attendue = echeance_du_classeur(ligne.a_planifier_texte)
        if not _memes_echeances(echeance, attendue):
            bilan.divergences.append(
                (
                    etiquette,
                    f"moteur {echeance.isoformat()} ≠ classeur "
                    f"« {ligne.a_planifier_texte} »",
                )
            )

        cle_planif = (employee_id, echeance.year, interview_type)
        if cle_planif in existants:
            continue
        existants.add(cle_planif)
        bilan.planifications.append(
            {
                "employee_id": employee_id,
                "company_id": company_id,
                "year": echeance.year,
                "status": "planifie",
                "planned_date": echeance.isoformat(),
                "interview_type": interview_type,
                "import_source": SOURCE,
            }
        )

    return bilan


def decider_refus_ecriture(
    url_base: Any, apply: bool, confirmation_production: bool
) -> Optional[str]:
    """Refuse une écriture en production sans confirmation explicite.

    Fonction pure, robuste à un `url_base` non converti (supabase-py rend un yarl.URL)
    et à None : ce garde-fou doit refuser ou autoriser, jamais lever.
    """
    if not apply:
        return None
    url_texte = "" if url_base is None else str(url_base)
    if PROJECT_REF_PRODUCTION in url_texte and not confirmation_production:
        return (
            f"Refus d'écrire : la base ciblée est la production ({url_texte}). "
            f"Ajouter {DRAPEAU_CONFIRMATION_PRODUCTION} pour écrire malgré tout."
        )
    return None


def verifier_pas_de_troncature(
    lignes: List[Dict[str, Any]], limite: int = LIMITE_POSTGREST
) -> None:
    """Une lecture tronquée ferait recréer des entretiens déjà en base."""
    if len(lignes) >= limite:
        raise SystemExit(
            f"Lecture des entretiens existants tronquée à {limite} lignes (limite "
            "PostgREST) : impossible de garantir l'absence de doublons. Paginer cette "
            "lecture avant de poursuivre."
        )


def _charger_base() -> Tuple[Any, Dict[str, str]]:
    import os

    from dotenv import load_dotenv
    from supabase import create_client

    load_dotenv(Path(__file__).resolve().parents[1] / ".env")
    client = create_client(
        os.environ["SUPABASE_URL"],
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ["SUPABASE_KEY"],
    )
    societes = {
        s["company_name"]: s["id"]
        for s in client.table("companies").select("id,company_name").execute().data
    }
    return client, societes


def analyser_arguments(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parseur = argparse.ArgumentParser(description=__doc__)
    parseur.add_argument("--fichier", type=Path, required=True)
    parseur.add_argument("--dry-run", action="store_true", default=True)
    parseur.add_argument("--apply", action="store_true")
    parseur.add_argument(
        "--confirmer-production",
        action="store_true",
        help=(
            "Requis en plus de --apply pour écrire sur la base de production "
            f"({PROJECT_REF_PRODUCTION})."
        ),
    )
    parseur.add_argument(
        "--ignorer-divergences",
        action="store_true",
        help="Écrire malgré un écart entre l'échéance calculée et celle du classeur.",
    )
    parseur.add_argument(
        "--aujourdhui",
        type=lambda v: datetime.strptime(v, "%Y-%m-%d").date(),
        default=None,
        help="Date de référence du calcul des échéances (défaut : aujourd'hui).",
    )
    return parseur.parse_args(argv)


def main() -> int:
    options = analyser_arguments()
    aujourdhui = options.aujourdhui or date.today()

    lignes = lire_classeur(options.fichier)
    reglages, bloques_reglage = deduire_reglages(lignes)

    client, societes = _charger_base()
    url_base = str(client.supabase_url)
    print(f"Base ciblée : {url_base}")

    refus = decider_refus_ecriture(
        url_base, options.apply, options.confirmer_production
    )
    if refus is not None:
        print(refus)
        return 1

    salaries = (
        client.table("employees")
        .select(
            "id,company_id,last_name,nom_usage,first_name,employment_status,"
            "statut,is_forfait_jour,hire_date"
        )
        .execute()
        .data
    )
    existants_bruts = (
        client.table("annual_reviews")
        .select("employee_id,year,interview_type")
        .execute()
        .data
    )
    verifier_pas_de_troncature(existants_bruts)
    existants = {
        (e["employee_id"], e["year"], e.get("interview_type") or "annual_performance")
        for e in existants_bruts
    }

    bilan = planifier(
        lignes, reglages, societes, salaries, existants, aujourdhui
    )
    bilan.bloques = bloques_reglage + bilan.bloques

    print(f"\n{len(lignes)} ligne(s) au classeur.")
    print(f"  {len(bilan.reglages)} réglage(s) de campagne à écrire")
    print(f"  {len(bilan.historiques)} entretien(s) passé(s) à reprendre")
    print(f"  {len(bilan.planifications)} entretien(s) à planifier")
    print(f"  {len(bilan.ignores)} ligne(s) ignorée(s)")
    print(f"  {len(bilan.divergences)} divergence(s) avec le classeur")
    print(f"  {len(bilan.bloques)} ligne(s) bloquée(s)")

    for etiquette, motif in bilan.ignores:
        print(f"  IGNORÉ      {etiquette} : {motif}")
    for etiquette, motif in bilan.divergences:
        print(f"  DIVERGENCE  {etiquette} : {motif}")
    for etiquette, motif in bilan.bloques:
        print(f"  BLOQUÉ      {etiquette} : {motif}")

    if not options.apply:
        print("\nMode simulation — rien n'a été écrit. Relancer avec --apply.")
        return 1 if bilan.bloques else 0

    if bilan.bloques:
        print("\nÉcriture refusée : résoudre d'abord les lignes bloquées.")
        return 1
    if bilan.divergences and not options.ignorer_divergences:
        print(
            "\nÉcriture refusée : le moteur et le classeur ne concordent pas. "
            "Relancer avec --ignorer-divergences pour écrire malgré tout."
        )
        return 1

    for company_id, valeurs in bilan.reglages.items():
        client.table("company_interview_settings").upsert(
            {**valeurs, "company_id": company_id}, on_conflict="company_id"
        ).execute()
    print(f"{len(bilan.reglages)} réglage(s) de campagne écrit(s).")

    for lot, libelle in (
        (bilan.historiques, "entretien(s) passé(s) repris"),
        (bilan.planifications, "entretien(s) planifié(s)"),
    ):
        for ligne in lot:
            client.table("annual_reviews").insert(ligne).execute()
        print(f"{len(lot)} {libelle}.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
