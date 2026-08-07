"""Charge les élus CSE depuis le classeur transmis par Elsa.

Le classeur ne contient pas les dates de mandat : la colonne « Date d'entrée » est la
date d'embauche du salarié, pas le mandat. Les dates viennent donc soit de deux colonnes
ajoutées au classeur, soit de --mandat (cas normal : un mandat par société, commun à
tous ses élus).

Aucun nom n'est écrit dans ce fichier : le dépôt est public, les données nominatives
restent sous data/.

--apply refuse d'écrire sur la base de production sans --confirmer-production : voir
`decider_refus_ecriture`.

Usage :
    python scripts/import_elus_cse.py --fichier <classeur.xlsx> --dry-run
    python scripts/import_elus_cse.py --fichier <classeur.xlsx> \\
        --mandat "CARTOL=2023-06-15:2027-06-14" --apply --confirmer-production
"""

from __future__ import annotations

import argparse
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

ROLES = {
    "Membre titulaire": "titulaire",
    "Membre suppléant": "suppleant",
    "Membre suppleant": "suppleant",
    "Secrétaire": "secretaire",
    "Secretaire": "secretaire",
    "Trésorier": "tresorier",
    "Tresorier": "tresorier",
}

# Le classeur nomme les sociétés autrement que la base.
SOCIETES = {
    "CARTOL": "Cartol Industrie",
    "LEWIS": "LEWIS",
    "MONT BLANC COMPOSITE (MBC)": "Mont Blanc Composite",
    "MBC": "Mont Blanc Composite",
    "COLORPLAST": "Colorplast",
    "COMITECH": "Comitech Composite",
    "MAJI": "MAJI",
    "ZONE 404": "Zone 404 Mars",
}

# Statuts éligibles à un mandat : un élu en préavis (en_sortie) ou pas encore intégré
# (en_onboarding) reste un élu. Cf. supabase/migrations/20260610140000_employees_status_
# allow_parti.sql (contrainte base : actif|active|en_sortie|en_onboarding|parti|inactif)
# et la convention `in ("actif", "active")` utilisée ailleurs dans le code.
STATUTS_ELIGIBLES = frozenset({"actif", "active", "en_sortie", "en_onboarding"})

# Référence du projet Supabase de production : --apply s'y refuse sans confirmation
# explicite (voir decider_refus_ecriture).
PROJECT_REF_PRODUCTION = "slleauhyjnmiawosvlcg"
DRAPEAU_CONFIRMATION_PRODUCTION = "--confirmer-production"

# Limite de lignes que PostgREST renvoie par requête sans pagination explicite : au-delà,
# une lecture est tronquée en silence.
LIMITE_POSTGREST = 1000


@dataclass
class LigneElu:
    societe: str
    nom: str
    prenom: str
    qualite: str
    college: Optional[str]
    debut_mandat: Optional[date]
    fin_mandat: Optional[date]


def cle_nom(valeur: Optional[str]) -> str:
    """Majuscules, sans accent, sans tiret ni espace — pour comparer deux noms."""
    sans_accent = unicodedata.normalize("NFD", (valeur or "").upper())
    sans_accent = sans_accent.encode("ascii", "ignore").decode()
    return sans_accent.replace("-", "").replace("'", "").replace(" ", "").strip()


def _vers_date(valeur: Any) -> Optional[date]:
    if valeur in (None, ""):
        return None
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    texte = str(valeur).strip()
    for motif in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(texte, motif).date()
        except ValueError:
            continue
    return None


def lire_classeur(chemin: Path) -> List[LigneElu]:
    """Lit le classeur d'Elsa. Colonnes de mandat facultatives."""
    import openpyxl

    ws = openpyxl.load_workbook(chemin, data_only=True).active
    entetes = [str(c.value or "").strip() for c in ws[1]]
    lignes: List[LigneElu] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        valeurs = dict(zip(entetes, row))
        if not valeurs.get("Nom"):
            continue
        lignes.append(
            LigneElu(
                societe=str(valeurs.get("Société") or "").strip(),
                nom=str(valeurs.get("Nom") or "").strip(),
                prenom=str(valeurs.get("Prénom") or "").strip(),
                qualite=str(valeurs.get("Qualité") or "").strip(),
                college=(str(valeurs.get("Collège")).strip()
                         if valeurs.get("Collège") not in (None, "", "Non précisé")
                         else None),
                debut_mandat=_vers_date(valeurs.get("Date début mandat")),
                fin_mandat=_vers_date(valeurs.get("Date fin mandat")),
            )
        )
    return lignes


def dates_mandat_par_societe(
    arguments: List[str],
) -> Dict[str, Tuple[date, date]]:
    """Convertit --mandat "SOCIETE=AAAA-MM-JJ:AAAA-MM-JJ" en table de correspondance."""
    resultat: Dict[str, Tuple[date, date]] = {}
    for argument in arguments or []:
        societe, _, periode = argument.partition("=")
        debut_texte, _, fin_texte = periode.partition(":")
        debut = _vers_date(debut_texte)
        fin = _vers_date(fin_texte)
        if debut is None or fin is None:
            raise SystemExit(
                f"--mandat illisible : {argument!r} "
                '(attendu "SOCIETE=AAAA-MM-JJ:AAAA-MM-JJ")'
            )
        if fin < debut:
            raise SystemExit(f"--mandat : la fin précède le début ({argument!r})")
        resultat[societe.strip().upper()] = (debut, fin)
    return resultat


def rapprocher(
    ligne: LigneElu, salaries: List[Dict[str, Any]]
) -> Optional[Dict[str, Any]]:
    """Retrouve le salarié correspondant, par nom de naissance ou nom d'usage."""
    cible = cle_nom(ligne.nom)
    candidats = [
        s
        for s in salaries
        if cle_nom(s.get("last_name")) == cible or cle_nom(s.get("nom_usage")) == cible
    ]
    if not candidats:
        return None
    if len(candidats) == 1:
        return candidats[0]
    prenom = cle_nom(ligne.prenom)
    exacts = [s for s in candidats if cle_nom(s.get("first_name")) == prenom]
    return exacts[0] if len(exacts) == 1 else None


def motif_blocage_statut(salarie: Dict[str, Any]) -> Optional[str]:
    """Renvoie un motif de blocage si le salarié rapproché n'est pas éligible.

    Éligibles : actif, active, en_sortie (préavis), en_onboarding — un élu en préavis
    reste un élu. Bloqués : parti, inactif, et tout statut inconnu ou absent :
    l'homonymie avec un salarié réellement actif reste possible.
    """
    statut = salarie.get("employment_status")
    statut_normalise = str(statut or "").strip().lower()
    if statut_normalise in STATUTS_ELIGIBLES:
        return None
    return f"salarié non actif ({statut or 'inconnu'})"


def analyser_arguments(argv: Optional[List[str]] = None) -> argparse.Namespace:
    """Construit et exécute le parseur de ligne de commande.

    Seul `options.apply` pilote le mode d'exécution (écriture vs simulation) — voir
    `main()`. `--dry-run` est accepté explicitement pour la lisibilité de la ligne de
    commande et pour empêcher --dry-run et --apply ensemble, mais ne porte aucune
    logique propre : lui donner un `default=True` recalculé après coup a longtemps fait
    croire qu'il pilotait quelque chose, alors que rien ne le lisait jamais.
    """
    parseur = argparse.ArgumentParser(description=__doc__)
    parseur.add_argument("--fichier", required=True, type=Path)
    parseur.add_argument("--mandat", action="append", default=[])
    groupe = parseur.add_mutually_exclusive_group()
    groupe.add_argument("--dry-run", action="store_true")
    groupe.add_argument("--apply", action="store_true")
    parseur.add_argument(
        "--confirmer-production",
        action="store_true",
        help=(
            "Requis en plus de --apply pour écrire sur la base de production "
            f"({PROJECT_REF_PRODUCTION})."
        ),
    )
    return parseur.parse_args(argv)


def decider_refus_ecriture(
    url_base: Any, apply: bool, confirmation_production: bool
) -> Optional[str]:
    """Décide si l'écriture doit être refusée. Fonction pure : ne lit ni le réseau ni
    l'environnement, prend l'URL de la base ciblée et les options, rend un message de
    refus ou None si l'écriture est autorisée.

    `url_base` accepte aussi bien une `str` qu'un objet convertible en chaîne (p. ex.
    l'objet `yarl.URL` renvoyé par `client.supabase_url` côté supabase-py) ou `None` :
    ce garde-fou protège une base de production, il ne doit jamais planter faute d'un
    type inattendu — il doit refuser ou autoriser, jamais lever une exception.

    En mode simulation (apply=False), l'écriture n'a jamais lieu : rien à refuser. En
    mode --apply sur la production, sans --confirmer-production, on refuse.
    """
    if not apply:
        return None
    url_texte = "" if url_base is None else str(url_base)
    if PROJECT_REF_PRODUCTION in url_texte and not confirmation_production:
        return (
            f"Refus d'écrire : la base ciblée est la production ({url_texte}). "
            f"Ajouter {DRAPEAU_CONFIRMATION_PRODUCTION} pour écrire malgré tout."
        )
    return None


def verifier_pas_de_troncature(
    lignes: List[Dict[str, Any]], limite: int = LIMITE_POSTGREST
) -> None:
    """Arrête le script si une lecture Supabase a pu être tronquée par PostgREST.

    Sans pagination, PostgREST plafonne une lecture à 1000 lignes. Pour les mandats déjà
    en base, une troncature est dangereuse : un mandat existant hors du jeu lu ⇒ le
    script le recrée en double.
    """
    if len(lignes) >= limite:
        raise SystemExit(
            f"Lecture des mandats existants tronquée à {limite} lignes (limite "
            "PostgREST) : impossible de garantir l'absence de doublons. Paginer cette "
            "lecture avant de poursuivre."
        )


def _charger_base() -> Tuple[Any, Dict[str, str]]:
    import os

    from dotenv import load_dotenv
    from supabase import create_client

    load_dotenv(Path(__file__).resolve().parents[1] / ".env")
    client = create_client(
        os.environ["SUPABASE_URL"],
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ["SUPABASE_KEY"],
    )
    societes = {
        s["company_name"]: s["id"]
        for s in client.table("companies").select("id,company_name").execute().data
    }
    return client, societes


def planifier(
    lignes: List[LigneElu],
    societes: Dict[str, str],
    salaries: List[Dict[str, Any]],
    existants: Set[Tuple[str, str]],
    mandats: Dict[str, Tuple[date, date]],
    aujourd_hui: Optional[date] = None,
) -> Tuple[List[Dict[str, Any]], List[Tuple[LigneElu, str]]]:
    """Détermine les mandats à créer et les lignes bloquées.

    `existants` est modifié en place : chaque mandat retenu y est ajouté au fil de la
    boucle, pour qu'une même ligne dupliquée dans le classeur ne crée pas deux mandats
    identiques dans le même run (elle prend alors la même branche qu'un mandat déjà en
    base).
    """
    aujourd_hui = aujourd_hui or date.today()
    a_creer: List[Dict[str, Any]] = []
    bloques: List[Tuple[LigneElu, str]] = []
    for ligne in lignes:
        nom_base = SOCIETES.get(ligne.societe.upper())
        company_id = societes.get(nom_base or "")
        debut, fin = ligne.debut_mandat, ligne.fin_mandat
        if debut is None or fin is None:
            depuis_option = mandats.get(ligne.societe.upper())
            if depuis_option:
                debut, fin = depuis_option
        role = ROLES.get(ligne.qualite)
        salarie = rapprocher(
            ligne, [s for s in salaries if s["company_id"] == company_id]
        )

        motif_statut = motif_blocage_statut(salarie) if salarie is not None else None

        if company_id is None:
            bloques.append((ligne, "société inconnue en base"))
        elif salarie is None:
            bloques.append((ligne, "aucun salarié rapproché"))
        elif motif_statut is not None:
            bloques.append((ligne, motif_statut))
        elif role is None:
            bloques.append((ligne, f"qualité inconnue : {ligne.qualite!r}"))
        elif debut is None or fin is None:
            bloques.append((ligne, "dates de mandat manquantes"))
        elif (salarie["id"], debut.isoformat()) in existants:
            print(f"  déjà en base  {ligne.nom} {ligne.prenom}")
        else:
            mandat = {
                "company_id": company_id,
                "employee_id": salarie["id"],
                "role": role,
                "college": ligne.college,
                "start_date": debut.isoformat(),
                "end_date": fin.isoformat(),
                # Un mandat déjà terminé ne doit jamais être importé comme actif : sinon
                # count_active_elected_members() (is_active seul) fait passer à tort la
                # société en « élu / conforme ».
                "is_active": fin >= aujourd_hui,
            }
            a_creer.append(mandat)
            existants.add((salarie["id"], debut.isoformat()))
            print(
                f"  à créer       {ligne.nom} {ligne.prenom} — {role} — "
                f"{debut:%d/%m/%Y} → {fin:%d/%m/%Y}"
            )

    return a_creer, bloques


def main() -> int:
    options = analyser_arguments()

    lignes = lire_classeur(options.fichier)
    mandats = dates_mandat_par_societe(options.mandat)
    client, societes = _charger_base()
    # client.supabase_url est un objet yarl.URL (supabase-py), pas une str : on le
    # convertit explicitement avant de le passer au garde-fou de production, qui reste
    # lui-même robuste à un objet non converti (voir decider_refus_ecriture).
    url_base = str(client.supabase_url)
    print(f"Base ciblée : {url_base}")

    refus = decider_refus_ecriture(
        url_base, options.apply, options.confirmer_production
    )
    if refus is not None:
        print(refus)
        return 1

    salaries = (
        client.table("employees")
        .select("id,company_id,last_name,nom_usage,first_name,employment_status")
        .execute()
        .data
    )
    existants_bruts = (
        client.table("cse_elected_members")
        .select("employee_id,start_date")
        .execute()
        .data
    )
    verifier_pas_de_troncature(existants_bruts)
    existants = {(e["employee_id"], e["start_date"]) for e in existants_bruts}

    a_creer, bloques = planifier(lignes, societes, salaries, existants, mandats)

    for ligne, motif in bloques:
        print(f"  BLOQUÉ        {ligne.societe} {ligne.nom} {ligne.prenom} : {motif}")

    print(f"\n{len(a_creer)} mandat(s) à créer, {len(bloques)} bloqué(s).")

    if not options.apply:
        print("Mode simulation — rien n'a été écrit. Relancer avec --apply pour écrire.")
        return 1 if bloques else 0

    if bloques:
        print("Écriture refusée : résoudre d'abord les lignes bloquées.")
        return 1

    for mandat in a_creer:
        client.table("cse_elected_members").insert(mandat).execute()
    print(f"{len(a_creer)} mandat(s) créé(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
