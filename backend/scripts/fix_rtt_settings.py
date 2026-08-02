"""Remise à plat des RTT (point #9).

Contexte
--------
Le moteur accordait 10 jours de RTT par an à toute entreprise non paramétrée, et
ne reconnaissait aucun salarié au forfait-jours. Les compteurs fantômes qui en
résultaient ont été neutralisés à la main, en posant dans
`employee_leave_adjustments.rtt_opening_balance` l'exact opposé du calcul.

Les deux bugs sont corrigés dans le moteur. Ce script remet la donnée d'aplomb :

1. il paramètre les entreprises qui ont des cadres au forfait-jours ;
2. il retire les contrepoids négatifs, devenus sans objet ;
3. il cale les soldes sur le calendrier de paie, seule source des jours déjà pris.

Les entreprises sans aucun forfait-jours restent volontairement sans ligne de
configuration : aucun paramétrage veut dire aucun RTT, ce qui est le droit.

Le forfait diffère d'une entreprise à l'autre — il vient de leur convention, pas
d'une règle générale. Les fichiers « jours de repos » des sociétés le donnent
noir sur blanc : Cartol est à 214 jours (13 jours de repos en 2026), Mont Blanc
Composite à 216 (11 jours). Ne pas généraliser une valeur observée ailleurs.

Usage
-----
    python scripts/fix_rtt_settings.py                    # simulation
    python scripts/fix_rtt_settings.py --apply            # écriture + sauvegarde
    python scripts/fix_rtt_settings.py --revert FICHIER   # restauration
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.database import supabase  # noqa: E402
from app.modules.absences.application.leave_settings_commands import (  # noqa: E402
    apply_rtt_solde_manual,
    update_employee_leave_adjustment,
    update_leave_settings,
)
from app.modules.absences.schemas.leave_settings import (  # noqa: E402
    EmployeeLeaveAdjustmentUpdate,
    LeaveSettingsUpdate,
)

YEAR = 2026
ACTIVE_STATUSES = ["actif", "en_sortie"]
DEFAULT_FORFAIT = 216

# Relevé dans le fichier « jours de repos » de chaque société. Toute entreprise
# absente de cette table prend DEFAULT_FORFAIT, valeur à confirmer sur pièce.
FORFAIT_BY_COMPANY = {
    "Cartol Industrie": 214,  # fichier CARTOL/RTT/Jour repos forfait 2026.xlsx
    "Mont Blanc Composite": 216,  # fichier MBC/jour repos cadre 2026.xlsx
}

# Calendrier de paie portant les soldes réels (ligne « Repos cadre », code .RE1).
MBC_CALENDAR = Path.home() / "Desktop" / "MBC" / "CALENDRIER 2026.xlsx"
MBC_SHEETS = [
    "BLONDEAU",
    "BORDELIER",
    "DROZ-VINCENT",
    "DULPHY",
    "GAILLET",
    "GILLET",
    "LABBE",
]

BACKUP_DIR = Path(__file__).resolve().parents[1] / "reports"


def _companies() -> dict[str, str]:
    rows = supabase.table("companies").select("id, company_name").execute().data
    return {r["id"]: r["company_name"] for r in rows}


def _employees() -> list[dict]:
    return (
        supabase.table("employees")
        .select("id, first_name, last_name, company_id, statut, is_forfait_jour")
        .in_("employment_status", ACTIVE_STATUSES)
        .execute()
        .data
    )


def _adjustments() -> list[dict]:
    return (
        supabase.table("employee_leave_adjustments")
        .select("employee_id, year, rtt_opening_balance")
        .eq("year", YEAR)
        .execute()
        .data
    )


def _leave_settings_rows() -> list[dict]:
    return supabase.table("company_leave_settings").select("*").execute().data


def read_mbc_soldes() -> dict[str, float]:
    """Solde de repos cadre au dernier mois renseigné du calendrier.

    Les mois à venir sont à zéro pour tout le monde ; on repère donc le dernier
    mois où quelqu'un porte une valeur, et on lit cette colonne pour chacun. Un
    zéro dans cette colonne est un vrai zéro — un cadre qui a tout consommé.
    """
    from openpyxl import load_workbook

    if not MBC_CALENDAR.exists():
        return {}
    wb = load_workbook(MBC_CALENDAR, data_only=True)
    series: dict[str, list[float]] = {}
    for name in MBC_SHEETS:
        if name not in wb.sheetnames:
            continue
        for row in wb[name].iter_rows(min_row=1, max_row=60, values_only=True):
            label = str(row[0]).strip().lower() if row[0] else ""
            if label.startswith("repos cadre"):
                series[name] = [c for c in row[1:] if isinstance(c, (int, float))]
    if not series:
        return {}
    width = max(len(v) for v in series.values())
    last = -1
    for i in range(width):
        if any(len(v) > i and v[i] > 0 for v in series.values()):
            last = i
    if last < 0:
        return {}
    return {n: float(v[last]) for n, v in series.items() if len(v) > last}


def do_revert(path: Path) -> int:
    snapshot = json.loads(path.read_text())
    print(f"Restauration depuis {path}\n")
    for row in snapshot["leave_settings"]:
        cid = row["company_id"]
        patch = LeaveSettingsUpdate(
            rtt_annual_days=row.get("rtt_annual_days"),
            rtt_use_forfait_jours_formula=bool(row.get("rtt_use_forfait_jours_formula")),
            rtt_forfait_annual_days=row.get("rtt_forfait_annual_days"),
            rtt_forfait_cadres_only=bool(row.get("rtt_forfait_cadres_only")),
        )
        update_leave_settings(cid, patch)
        print(f"  config restaurée : {cid}")
    employees = {e["id"]: e for e in _employees()}
    restored = 0
    for row in snapshot["adjustments"]:
        emp = employees.get(row["employee_id"])
        if not emp:
            continue
        update_employee_leave_adjustment(
            emp["company_id"],
            emp["id"],
            row["year"],
            EmployeeLeaveAdjustmentUpdate(
                rtt_opening_balance=row.get("rtt_opening_balance") or 0
            ),
        )
        restored += 1
    print(f"  {restored} ajustements restaurés")
    print("\n⚠ Les lignes de configuration créées par --apply ne sont pas supprimées,")
    print("  seulement remises à leurs valeurs d'avant.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="écrire en base")
    parser.add_argument("--revert", metavar="FICHIER", help="restaurer une sauvegarde")
    args = parser.parse_args()

    if args.revert:
        return do_revert(Path(args.revert))

    companies = _companies()
    employees = _employees()
    by_employee = {e["id"]: e for e in employees}

    forfait_employees: dict[str, list[dict]] = {}
    for e in employees:
        if e["is_forfait_jour"]:
            forfait_employees.setdefault(e["company_id"], []).append(e)

    to_configure = sorted(forfait_employees, key=lambda c: companies.get(c, ""))
    untouched = sorted(set(companies) - set(to_configure), key=lambda c: companies[c])

    print(f"Mode : {'ÉCRITURE' if args.apply else 'simulation'}\n")

    print("1. Paramétrage")
    for cid in to_configure:
        name = companies[cid]
        forfait = FORFAIT_BY_COMPANY.get(name, DEFAULT_FORFAIT)
        origin = "relevé" if name in FORFAIT_BY_COMPANY else "À CONFIRMER"
        print(
            f"   {name:<24} forfait {forfait} j ({origin})"
            f"  {len(forfait_employees[cid])} salarié(s)"
        )
    print("\n   Sans configuration, donc sans RTT :")
    for cid in untouched:
        print(f"   {companies[cid]}")

    adjustments = _adjustments()
    negative = [a for a in adjustments if (a.get("rtt_opening_balance") or 0) < 0]
    print(f"\n2. Contrepoids négatifs à retirer : {len(negative)}")

    soldes = read_mbc_soldes()
    print(f"\n3. Soldes réels lus au calendrier MBC : {len(soldes)}")
    for name, value in sorted(soldes.items()):
        print(f"   {name:<16}{value:>6}")
    if not soldes:
        print("   (calendrier introuvable — soldes laissés au quota théorique)")

    if not args.apply:
        print("\nSimulation terminée — relancer avec --apply.")
        return 0

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = BACKUP_DIR / f"rtt_backup_{stamp}.json"
    backup.write_text(
        json.dumps(
            {"leave_settings": _leave_settings_rows(), "adjustments": adjustments},
            indent=2,
            default=str,
        )
    )
    print(f"\nSauvegarde écrite : {backup}")

    print("\nParamétrage…")
    for cid in to_configure:
        name = companies[cid]
        res = update_leave_settings(
            cid,
            LeaveSettingsUpdate(
                rtt_use_forfait_jours_formula=True,
                rtt_forfait_annual_days=FORFAIT_BY_COMPANY.get(name, DEFAULT_FORFAIT),
                rtt_forfait_cadres_only=True,
                rtt_forfait_cp_ouvres_deduction=25,
            ),
        )
        print(f"   {name:<24} forfait={res.rtt_forfait_annual_days}")

    print("\nRetrait des contrepoids…")
    done = skipped = 0
    for adj in negative:
        emp = by_employee.get(adj["employee_id"])
        if not emp:
            skipped += 1
            continue
        update_employee_leave_adjustment(
            emp["company_id"],
            emp["id"],
            adj["year"],
            EmployeeLeaveAdjustmentUpdate(rtt_opening_balance=0),
        )
        done += 1
    print(f"   {done} remis à zéro, {skipped} ignorés (salarié inactif)")

    if soldes:
        print("\nCalage des soldes sur le calendrier…")
        by_lastname: dict[str, dict] = {}
        for e in employees:
            by_lastname.setdefault(e["last_name"].strip().upper(), e)
        for sheet, value in sorted(soldes.items()):
            emp = by_lastname.get(sheet.strip().upper())
            if not emp:
                print(f"   {sheet:<16} salarié introuvable, ignoré")
                continue
            apply_rtt_solde_manual(
                emp["company_id"],
                emp["id"],
                YEAR,
                rtt_solde=value,
                note=f"Solde repos cadre relevé au calendrier de paie {YEAR}",
            )
            print(f"   {sheet:<16} solde posé à {value}")

    print(f"\nTerminé. Retour arrière : --revert {backup}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
